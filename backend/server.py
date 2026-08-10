"""
JONAH ENTERPRISES - Door Installation Work Management System
FastAPI + MongoDB backend with JWT auth, role-based access, GPS check-in/out,
work reports with photo evidence, admin approval workflow, payment ledger.
"""

import os
import io
import asyncio
import base64
import binascii
import math
import re
import secrets
import uuid
import logging
from datetime import datetime, timedelta, timezone
from enum import Enum
from pathlib import Path
from typing import Annotated, List, Optional
from urllib.parse import urlencode, quote

import certifi
import httpx
import jwt
from xml.sax.saxutils import escape as _xml_escape
from dotenv import load_dotenv
from fastapi import FastAPI, APIRouter, Depends, HTTPException, Query, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse, RedirectResponse
from fastapi.security import OAuth2PasswordBearer
from motor.motor_asyncio import AsyncIOMotorClient
from passlib.context import CryptContext
from pydantic import BaseModel, Field
from PIL import Image, ImageDraw, ImageFont, UnidentifiedImageError

import excel_sync
from crypto_utils import encrypt_secret, decrypt_secret

# --- Setup ---------------------------------------------------------------
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

JWT_SECRET = os.environ.get("JWT_SECRET")
if not JWT_SECRET:
    # Fail loudly instead of silently minting a per-process random secret — a random
    # fallback would invalidate every outstanding token on each restart and make
    # replicas reject each other's tokens, which is easy to miss until it happens.
    raise RuntimeError("JWT_SECRET environment variable must be set (see backend/.env)")
JWT_ALGO = "HS256"
ACCESS_MINUTES = 60 * 24 * 30  # 30 days for field workers

PUSH_BASE_URL = "https://integrations.emergentagent.com"
PUSH_KEY = os.environ.get("EMERGENT_PUSH_KEY", "placeholder")

# Explicit allow-list — never combine allow_origins=["*"] with allow_credentials=True,
# Starlette echoes back the request Origin in that combo, effectively allowing any site.
CORS_ORIGINS = [o.strip() for o in os.environ.get(
    "CORS_ORIGINS", "http://localhost:3000,http://localhost:5173,http://localhost:19006"
).split(",") if o.strip()]

# Obviously-fake demo workers/sites/reports — never seed these against a real deployment.
SEED_DEMO_DATA = os.environ.get("SEED_DEMO_DATA", "true").lower() == "true"

_mongo_url = os.environ["MONGO_URL"]
# Passing tlsCAFile implicitly forces TLS in PyMongo even for a plain mongodb:// URL
# with no tls param — fine (required, even) for Atlas's mongodb+srv:// URLs, but it
# silently breaks a local/self-hosted mongod that isn't TLS-configured. Only apply it
# for +srv connection strings, which is the standard hallmark of an Atlas connection.
_mongo_kwargs = {"tlsCAFile": certifi.where()} if _mongo_url.startswith("mongodb+srv://") else {}
client = AsyncIOMotorClient(_mongo_url, **_mongo_kwargs)
db = client[os.environ["DB_NAME"]]

pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto")
DUMMY_HASH = pwd_ctx.hash("dummy-timing-hash")
oauth2 = OAuth2PasswordBearer(tokenUrl="/api/auth/login", auto_error=False)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("jonah")

app = FastAPI(title="JONAH Enterprises WMS")
api = APIRouter(prefix="/api")


# --- Models --------------------------------------------------------------
class Role(str, Enum):
    worker = "Worker"
    admin = "Admin"
    client = "Client"


class LeaveStatus(str, Enum):
    pending = "Pending"
    approved = "Approved"
    rejected = "Rejected"


class PaymentRequestStatus(str, Enum):
    pending = "Pending"
    approved = "Approved"
    rejected = "Rejected"
    paid = "Paid"


class ExcelMatchStatus(str, Enum):
    unresolved = "Unresolved"   # never checked against Excel yet
    matched = "Matched"         # linked to an existing Excel master row
    create_pending = "Create Pending"  # admin confirmed: no Excel equivalent, safe to auto-create
    created = "Created"         # the app has created the Excel row for this record


class ExcelSyncStatus(str, Enum):
    not_applicable = "Not Applicable"  # Excel Integration isn't configured/connected
    pending = "Pending"
    synced = "Synced"
    failed = "Failed"


class ApprovalStatus(str, Enum):
    pending = "Pending"
    approved = "Approved"
    rejected = "Rejected"
    correction = "Correction"


class LoginIn(BaseModel):
    username: str
    password: str


class ChangePwdIn(BaseModel):
    current_password: str
    new_password: str = Field(min_length=8)


class ForgotPwdIn(BaseModel):
    username: str


class UserOut(BaseModel):
    id: str
    username: str
    role: Role
    name: Optional[str] = None
    worker_id: Optional[str] = None


class TokenOut(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserOut


class WorkerIn(BaseModel):
    employee_id: str
    name: str
    mobile: str
    username: str
    password: str = Field(min_length=8)
    joining_date: Optional[str] = None
    default_rate: float = Field(default=250.0, gt=0)
    notes: Optional[str] = ""


class WorkerUpdate(BaseModel):
    name: Optional[str] = None
    mobile: Optional[str] = None
    default_rate: Optional[float] = Field(default=None, gt=0)
    status: Optional[str] = None
    notes: Optional[str] = None


class SiteIn(BaseModel):
    site_name: str
    client_name: str
    address: str
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    contact_person: Optional[str] = ""
    contact_number: Optional[str] = ""
    status: str = "Active"
    notes: Optional[str] = ""
    door_type_rates: dict = {}  # {door_type_id: rate} — per-site override of a door type's default rate
    target_doors: dict = {}  # {door_type_id: assigned_count} — omitted key = no ceiling for that type
    assigned_worker_ids: List[str] = []  # empty = open to every worker (unchanged legacy behavior)
    geofence_radius_m: Optional[int] = None  # unset = fall back to the global default


class DoorTypeIn(BaseModel):
    name: str
    default_rate: float = Field(gt=0)
    status: str = "Active"


class CheckInIn(BaseModel):
    site_id: str
    latitude: float
    longitude: float
    accuracy: Optional[float] = None
    client_sync_id: Optional[str] = None  # idempotency


class CheckOutIn(BaseModel):
    session_id: str
    latitude: float
    longitude: float
    accuracy: Optional[float] = None
    work_completed: bool = True
    remarks: Optional[str] = ""


class DoorLineItem(BaseModel):
    door_type_id: str
    count: int = Field(gt=0)


class WorkReportIn(BaseModel):
    session_id: str
    doors: List[DoorLineItem] = Field(min_length=1)  # one entry per door type installed
    notes: Optional[str] = ""
    work_completed: bool = True
    before_photos: List[str] = []  # base64 strings
    after_photos: List[str] = []
    materials: List[dict] = []  # [{name, qty, unit, unit_cost}]
    client_sync_id: Optional[str] = None


class WorkReportUpdateIn(BaseModel):
    # Self-edit is scoped to counts/notes only, not photos — re-watermarking an
    # already-watermarked photo would stack a second strip on it, so photo edits
    # are deliberately out of scope here (withdraw and resubmit fresh instead).
    doors: List[DoorLineItem] = Field(min_length=1)
    notes: Optional[str] = ""
    work_completed: bool = True


class ApprovalIn(BaseModel):
    remarks: Optional[str] = ""
    edited_door_count: Optional[int] = None
    edited_rate: Optional[float] = None


class PaymentIn(BaseModel):
    worker_id: str
    amount: float = Field(gt=0)
    payment_date: Optional[str] = None
    payment_method: str = "Cash"
    transaction_reference: Optional[str] = ""
    notes: Optional[str] = ""
    client_sync_id: Optional[str] = None  # idempotency — a double-click must not double-pay


class PaymentRequestIn(BaseModel):
    amount: float = Field(gt=0)
    note: Optional[str] = ""


class PaymentRequestDecision(BaseModel):
    remarks: Optional[str] = ""


class PaymentRequestPayIn(BaseModel):
    amount: Optional[float] = Field(default=None, gt=0)  # defaults to the requested amount if unset
    payment_method: str = "UPI"
    transaction_reference: Optional[str] = ""
    notes: Optional[str] = ""
    client_sync_id: Optional[str] = None


class SettingsIn(BaseModel):
    company_name: Optional[str] = None
    address: Optional[str] = None
    mobile: Optional[str] = None
    email: Optional[str] = None
    currency: Optional[str] = None
    default_rate: Optional[float] = Field(default=None, gt=0)
    logo_base64: Optional[str] = None
    shift_start_time: Optional[str] = None  # "HH:MM", IST — unset disables late flagging
    shift_end_time: Optional[str] = None  # "HH:MM", IST — unset disables early-leave flagging
    late_grace_minutes: Optional[int] = None
    early_leave_grace_minutes: Optional[int] = None


class ExcelIntegrationSettingsIn(BaseModel):
    tenant_id: str = "common"
    client_id: str
    client_secret: str
    workbook_file_path: str  # e.g. "/Documents/JONAH ENTERPRISES-ERP-2026-FINAL.xlsm"
    redirect_uri: str


class BankingSettingsIn(BaseModel):
    provider_name: str = ""
    api_base_url: str = ""
    api_key: Optional[str] = ""
    client_id: Optional[str] = ""
    client_secret: Optional[str] = ""
    merchant_id: Optional[str] = ""
    webhook_url: Optional[str] = ""
    webhook_secret: Optional[str] = ""
    environment: str = "Sandbox"  # Sandbox | Production
    enabled: bool = False


class ExcelMappingConfirmIn(BaseModel):
    excel_id: Optional[str] = None  # link to this existing Excel row's ID
    create_new: bool = False        # confirm this app record has no Excel equivalent


# --- Helpers -------------------------------------------------------------
def now_utc():
    return datetime.now(timezone.utc)


def issue_token(user_doc: dict) -> str:
    payload = {
        "sub": user_doc["id"],
        "username": user_doc["username"],
        "role": user_doc["role"],
        "tv": user_doc.get("token_version", 0),
        "iat": now_utc(),
        "exp": now_utc() + timedelta(minutes=ACCESS_MINUTES),
        "jti": secrets.token_hex(8),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)


async def get_user(token: Annotated[Optional[str], Depends(oauth2)]) -> dict:
    unauth = HTTPException(status_code=401, detail="Not authenticated")
    if not token:
        raise unauth
    try:
        p = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
        u = await db.users.find_one({"id": p["sub"], "disabled": {"$ne": True}}, {"_id": 0})
        if not u:
            raise unauth
        # Tokens issued before a password change carry a stale "tv" and are rejected,
        # so a leaked token stops working as soon as the user changes their password.
        if p.get("tv", 0) != u.get("token_version", 0):
            raise unauth
        return u
    except jwt.PyJWTError:
        raise unauth


def require_role(*roles: Role):
    async def _guard(u=Depends(get_user)):
        if u["role"] not in [r.value for r in roles]:
            raise HTTPException(status_code=403, detail="Forbidden")
        return u
    return _guard


def public_user(u: dict) -> UserOut:
    return UserOut(
        id=u["id"],
        username=u["username"],
        role=u["role"],
        name=u.get("name"),
        worker_id=u.get("worker_id"),
    )


async def get_door_type_rate(door_type_id: str, site_id: str) -> tuple:
    """Rate for a door type at a given site: site-specific override if set, else the
    door type's own default_rate. Returns (name, rate)."""
    door_type = await db.door_types.find_one({"id": door_type_id}, {"_id": 0})
    if not door_type:
        raise HTTPException(400, f"Unknown door type: {door_type_id}")
    site = await db.sites.find_one({"id": site_id}, {"_id": 0})
    override = (site or {}).get("door_type_rates", {}).get(door_type_id)
    rate = float(override) if override is not None else float(door_type["default_rate"])
    return door_type["name"], rate


async def get_effective_rate(worker_id: str, site_id: str) -> float:
    """Legacy flat-rate fallback — used only where no door type applies (e.g. admin
    manually editing a report's total at approval time without changing its breakdown)."""
    worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    if worker and worker.get("default_rate"):
        return float(worker["default_rate"])
    settings = await db.settings.find_one({"id": "global"}, {"_id": 0})
    if settings and settings.get("default_rate"):
        return float(settings["default_rate"])
    return 250.0


GEOFENCE_WARNING_METERS = 500  # default when a site has no geofence_radius_m of its own
IST = timezone(timedelta(hours=5, minutes=30))  # the business operates in India only


def minutes_since_midnight_ist(iso_str: str) -> int:
    dt = datetime.fromisoformat(iso_str)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    local = dt.astimezone(IST)
    return local.hour * 60 + local.minute


def parse_hm_to_minutes(hm: str) -> int:
    h, m = hm.split(":")
    return int(h) * 60 + int(m)


def haversine_meters(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    r = 6371000  # Earth radius, meters
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lng2 - lng1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


async def get_installed_count(site_id: str, door_type_id: str, exclude_report_id: Optional[str] = None) -> int:
    """Sum of this door type already recorded at this site across every report that
    isn't Rejected — Pending/Correction count too, so the ceiling can't be worked
    around by racing multiple pending submissions ahead of admin review.
    exclude_report_id lets an edit re-check the ceiling without double-counting
    the very report being edited against itself."""
    total = 0
    async for r in db.work_reports.find(
        {"site_id": site_id, "approval_status": {"$ne": ApprovalStatus.rejected.value}},
        {"_id": 0, "id": 1, "door_breakdown": 1},
    ):
        if exclude_report_id and r.get("id") == exclude_report_id:
            continue
        for item in r.get("door_breakdown", []):
            if item.get("door_type_id") == door_type_id:
                total += item.get("count", 0)
    return total


async def add_notification(user_id: str, title: str, message: str, action_url: Optional[str] = None):
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "title": title,
        "message": message,
        "is_read": False,
        "created_at": now_utc().isoformat(),
    })
    # fire push (non-blocking)
    try:
        data = {"title": title, "message": message}
        if action_url:
            data["action_url"] = action_url
        await send_push([user_id], data)
    except Exception as e:
        logger.warning(f"push send failed (non-blocking): {e}")


async def audit(user: dict, action: str, entity_type: str, entity_id: str, old=None, new=None):
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "user_name": user.get("name") or user.get("username"),
        "action": action,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "old_data": old,
        "new_data": new,
        "created_at": now_utc().isoformat(),
    })


def clean(doc: dict) -> dict:
    if doc and "_id" in doc:
        doc.pop("_id", None)
    return doc


# --- Photo watermark -----------------------------------------------------
MAX_PHOTO_B64_CHARS = 12_000_000  # ~9MB decoded — generous ceiling for a phone photo


def _photo_too_large(b64_data_uri: str) -> bool:
    raw = b64_data_uri.split(",", 1)[1] if "," in b64_data_uri else b64_data_uri
    return len(raw) > MAX_PHOTO_B64_CHARS


ALLOWED_PHOTO_FORMATS = {"JPEG", "PNG", "WEBP"}


def _validate_photo(b64_data_uri: str, label: str) -> None:
    """Re-validates on the server that a claimed 'photo' is actually a decodable
    image in an allowed format, before it's trusted as evidence. Without this,
    _watermark_photo's broad except-and-fall-back-to-original would silently
    accept and store arbitrary non-image bytes a worker (or a compromised
    client) submitted, unwatermarked and untagged."""
    try:
        raw = b64_data_uri.split(",", 1)[1] if "," in b64_data_uri else b64_data_uri
        decoded = base64.b64decode(raw, validate=True)
        img = Image.open(io.BytesIO(decoded))
        img.verify()  # structural check only — img is unusable after this, which is fine, we only needed the check
        # Re-open a fresh instance: verify() invalidates the one above but the format
        # attribute set during the initial open() is still what we need to check.
        fmt = img.format
    except (UnidentifiedImageError, Image.DecompressionBombError, binascii.Error, ValueError, OSError):
        raise HTTPException(400, f"{label} is not a valid image. Please retake the photo and try again.")
    if fmt not in ALLOWED_PHOTO_FORMATS:
        raise HTTPException(400, f"{label} is a {fmt or 'unknown'} file — only JPEG, PNG, or WEBP photos are accepted.")


def _watermark_photo(b64_data_uri: str, worker_name: str, site_name: str,
                     lat: Optional[float], lng: Optional[float],
                     company: str = "JONAH ENTERPRISES") -> str:
    """Adds a semi-transparent brown/gold info strip to the bottom of a photo.
    Returns the watermarked photo as a base64 data URI (JPEG). If anything
    fails we return the original untouched so we never lose the worker's
    evidence."""
    try:
        # Parse "data:image/...;base64,XXXX" or bare base64
        raw = b64_data_uri.split(",", 1)[1] if "," in b64_data_uri else b64_data_uri
        img = Image.open(io.BytesIO(base64.b64decode(raw))).convert("RGB")
        w, h = img.size
        # Downscale huge photos (max width 1600) to keep documents small
        if w > 1600:
            ratio = 1600 / w
            img = img.resize((1600, int(h * ratio)))
            w, h = img.size

        strip_h = max(90, int(h * 0.14))
        strip = Image.new("RGBA", (w, strip_h), (43, 32, 25, 210))  # brown 82% alpha
        overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
        overlay.paste(strip, (0, h - strip_h))
        draw = ImageDraw.Draw(overlay)

        try:
            font_lg = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", int(strip_h * 0.22))
            font_sm = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", int(strip_h * 0.15))
        except Exception:
            font_lg = ImageFont.load_default()
            font_sm = ImageFont.load_default()

        ts = now_utc().strftime("%d/%m/%Y %H:%M UTC")
        gps_txt = f"GPS: {lat:.5f}, {lng:.5f}" if lat is not None and lng is not None else "GPS: unavailable"
        pad = int(strip_h * 0.12)
        y0 = h - strip_h + pad
        draw.text((pad, y0), company, fill=(192, 152, 54, 255), font=font_lg)  # gold
        line2_y = y0 + int(strip_h * 0.30)
        draw.text((pad, line2_y), f"Site: {site_name}   Worker: {worker_name}",
                  fill=(245, 235, 225, 255), font=font_sm)
        line3_y = line2_y + int(strip_h * 0.22)
        draw.text((pad, line3_y), f"{ts}   {gps_txt}",
                  fill=(245, 235, 225, 255), font=font_sm)

        merged = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
        out = io.BytesIO()
        merged.save(out, format="JPEG", quality=75, optimize=True)
        return "data:image/jpeg;base64," + base64.b64encode(out.getvalue()).decode()
    except Exception as e:
        logger.warning(f"Watermark failed, keeping original: {e}")
        return b64_data_uri


# --- Push notifications (Emergent managed) -------------------------------
_push_client = httpx.AsyncClient(
    base_url=PUSH_BASE_URL,
    headers={"X-Push-Key": PUSH_KEY},
    timeout=10.0,
)


async def send_push(recipients: List[str], data: dict, idempotency_key: Optional[str] = None) -> None:
    """Relay push notifications through the Emergent managed service.
    Never blocks the primary operation — callers should still catch."""
    if not recipients:
        return
    if "title" not in data or "message" not in data:
        raise ValueError("push data must include title and message")
    # chunk to 100 per playbook
    for i in range(0, len(recipients), 100):
        chunk = recipients[i:i + 100]
        payload = {"recipients": chunk, "data": data}
        if idempotency_key:
            payload["$idempotency_key"] = f"{idempotency_key}-{i}"
        try:
            resp = await _push_client.post("/api/v1/push/trigger", json=payload)
            if resp.status_code >= 400:
                logger.warning(f"Push relay non-2xx: {resp.status_code} {resp.text[:200]}")
        except Exception as e:
            logger.warning(f"Push relay error (non-blocking): {e}")


class RegisterPushIn(BaseModel):
    user_id: str
    platform: str
    device_token: str


# --- Auth endpoints ------------------------------------------------------
# Login throttle backed by Mongo (db.login_attempts) rather than an in-process
# dict — a dict resets on every restart and, more importantly, doesn't share
# state across multiple Railway instances/replicas, so an attacker spread
# across instances would get 5 tries per instance instead of 5 total.
LOGIN_MAX_ATTEMPTS = 5
LOGIN_WINDOW_SECONDS = 15 * 60


async def _login_rate_limited(key: str) -> bool:
    cutoff = now_utc() - timedelta(seconds=LOGIN_WINDOW_SECONDS)
    count = await db.login_attempts.count_documents({"username": key, "created_at": {"$gte": cutoff}})
    return count >= LOGIN_MAX_ATTEMPTS


async def _record_login_failure(key: str):
    await db.login_attempts.insert_one({"username": key, "created_at": now_utc()})


@api.post("/auth/login", response_model=TokenOut)
async def login(body: LoginIn):
    key = body.username.lower().strip()
    if await _login_rate_limited(key):
        raise HTTPException(status_code=429, detail="Too many login attempts. Try again later.")
    u = await db.users.find_one({"username": key}, {"_id": 0})
    valid = pwd_ctx.verify(body.password, u["password_hash"]) if u else pwd_ctx.verify(body.password, DUMMY_HASH)
    if not u or not valid or u.get("disabled"):
        await _record_login_failure(key)
        raise HTTPException(status_code=401, detail="Incorrect username or password")
    await db.login_attempts.delete_many({"username": key})
    return TokenOut(access_token=issue_token(u), user=public_user(u))


@api.get("/auth/me", response_model=UserOut)
async def me(u=Depends(get_user)):
    return public_user(u)


@api.post("/auth/change-password")
async def change_pwd(body: ChangePwdIn, u=Depends(get_user)):
    if not pwd_ctx.verify(body.current_password, u["password_hash"]):
        raise HTTPException(status_code=400, detail="Current password incorrect")
    await db.users.update_one(
        {"id": u["id"]},
        {"$set": {"password_hash": pwd_ctx.hash(body.new_password)}, "$inc": {"token_version": 1}}
    )
    return {"ok": True}


@api.post("/auth/forgot-password")
async def forgot(body: ForgotPwdIn):
    # In dev: reset to default password "worker123" if user exists (admin-controlled).
    # Real prod would send email. We keep response identical.
    u = await db.users.find_one({"username": body.username.lower().strip()})
    if u:
        logger.info(f"Password reset requested for {body.username}")
    return {"message": "If the account exists, contact your administrator for a password reset."}


# --- Sites ---------------------------------------------------------------
@api.get("/sites")
async def list_sites(u=Depends(get_user), status_f: Optional[str] = Query(None, alias="status")):
    q = {}
    if status_f:
        q["status"] = status_f
    elif u["role"] == Role.worker.value:
        q["status"] = "Active"
    if u["role"] == Role.worker.value:
        # A site with no team set is open to everyone (legacy behavior for existing
        # sites); once an admin picks a team, only those workers see/can check into it.
        q["$or"] = [
            {"assigned_worker_ids": {"$exists": False}},
            {"assigned_worker_ids": {"$size": 0}},
            {"assigned_worker_ids": u.get("worker_id")},
        ]
    if u["role"] == Role.client.value:
        # Clients only ever see sites they're assigned to, regardless of ?status=
        q["id"] = {"$in": u.get("site_ids") or []}
    items = await db.sites.find(q, {"_id": 0}).to_list(1000)
    return items


@api.post("/sites")
async def create_site(body: SiteIn, u=Depends(require_role(Role.admin))):
    doc = body.dict()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = now_utc().isoformat()
    doc["excel_site_id"] = None
    doc["excel_match_status"] = ExcelMatchStatus.unresolved.value
    await db.sites.insert_one(doc.copy())
    await audit(u, "create", "site", doc["id"], None, doc)
    return clean(doc)


@api.put("/sites/{site_id}")
async def update_site(site_id: str, body: SiteIn, u=Depends(require_role(Role.admin))):
    old = await db.sites.find_one({"id": site_id}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Site not found")
    upd = body.dict()
    await db.sites.update_one({"id": site_id}, {"$set": upd})
    await audit(u, "update", "site", site_id, old, upd)
    return {**old, **upd}


@api.delete("/sites/{site_id}")
async def delete_site(site_id: str, u=Depends(require_role(Role.admin))):
    await db.sites.update_one({"id": site_id}, {"$set": {"status": "Disabled"}})
    await audit(u, "disable", "site", site_id)
    return {"ok": True}


# --- Door Types ------------------------------------------------------------
@api.get("/door-types")
async def list_door_types(u=Depends(get_user), status_f: Optional[str] = Query(None, alias="status")):
    q = {}
    if status_f:
        q["status"] = status_f
    elif u["role"] == Role.worker.value:
        q["status"] = "Active"  # workers only pick from active types when submitting reports
    return await db.door_types.find(q, {"_id": 0}).sort("name", 1).to_list(200)


@api.post("/door-types")
async def create_door_type(body: DoorTypeIn, u=Depends(require_role(Role.admin))):
    existing = await db.door_types.find_one({"name": {"$regex": f"^{re.escape(body.name.strip())}$", "$options": "i"}})
    if existing:
        raise HTTPException(400, f'A door type named "{existing["name"]}" already exists')
    doc = body.dict()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = now_utc().isoformat()
    doc["excel_item_id"] = None
    doc["excel_match_status"] = ExcelMatchStatus.unresolved.value
    # cached from the matched M_ITEM row once mapped — Excel's Category/Specification
    # split doesn't exist in the app's flat door_type name, so we mirror Excel's own
    # values rather than trying to derive them from the app's name string.
    doc["excel_work_category"] = None
    doc["excel_specification"] = None
    await db.door_types.insert_one(doc.copy())
    await audit(u, "create", "door_type", doc["id"], None, doc)
    return clean(doc)


@api.put("/door-types/{door_type_id}")
async def update_door_type(door_type_id: str, body: DoorTypeIn, u=Depends(require_role(Role.admin))):
    old = await db.door_types.find_one({"id": door_type_id}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Door type not found")
    dupe = await db.door_types.find_one({
        "id": {"$ne": door_type_id},
        "name": {"$regex": f"^{re.escape(body.name.strip())}$", "$options": "i"},
    })
    if dupe:
        raise HTTPException(400, f'A door type named "{dupe["name"]}" already exists')
    upd = body.dict()
    await db.door_types.update_one({"id": door_type_id}, {"$set": upd})
    await audit(u, "update", "door_type", door_type_id, old, upd)
    return {**old, **upd}


@api.get("/sites/{site_id}/progress")
async def site_door_progress(site_id: str, u=Depends(get_user)):
    """Per-door-type target/installed/remaining at this site, for door types that
    have a target set. Lets the report form show remaining quantity before a
    worker submits, instead of only rejecting after the fact."""
    site = await db.sites.find_one({"id": site_id}, {"_id": 0})
    if not site:
        raise HTTPException(404, "Site not found")
    targets = site.get("target_doors", {}) or {}
    door_types = await db.door_types.find({"status": "Active"}, {"_id": 0}).to_list(200)
    rows = []
    for dt in door_types:
        target = targets.get(dt["id"])
        installed = await get_installed_count(site_id, dt["id"]) if target is not None else None
        rows.append({
            "door_type_id": dt["id"],
            "door_type_name": dt["name"],
            "target": target,
            "installed": installed,
            "remaining": max(target - installed, 0) if target is not None else None,
        })
    return rows


# --- Workers -------------------------------------------------------------
@api.get("/workers")
async def list_workers(u=Depends(require_role(Role.admin))):
    items = await db.workers.find({}, {"_id": 0}).to_list(1000)
    return items


@api.post("/workers")
async def create_worker(body: WorkerIn, u=Depends(require_role(Role.admin))):
    if await db.users.find_one({"username": body.username.lower().strip()}):
        raise HTTPException(400, "Username already exists")
    worker_id = str(uuid.uuid4())
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "username": body.username.lower().strip(),
        "password_hash": pwd_ctx.hash(body.password),
        "role": Role.worker.value,
        "name": body.name,
        "worker_id": worker_id,
        "disabled": False,
        "created_at": now_utc().isoformat(),
    }
    worker_doc = {
        "id": worker_id,
        "user_id": user_id,
        "employee_id": body.employee_id,
        "name": body.name,
        "mobile": body.mobile,
        "username": body.username.lower().strip(),
        "joining_date": body.joining_date or now_utc().date().isoformat(),
        "default_rate": body.default_rate,
        "status": "Active",
        "notes": body.notes or "",
        "created_at": now_utc().isoformat(),
        # Excel Integration — resolved by the one-time Excel Master Mapping screen
        "excel_party_id": None,
        "excel_match_status": ExcelMatchStatus.unresolved.value,
        # Per-worker automation — set the moment this worker's first valid app
        # work entry is created (not on login). Independent per worker.
        "automation_enabled": False,
        "first_valid_entry_at": None,
        "automation_started_at": None,
    }
    await db.users.insert_one(user_doc.copy())
    await db.workers.insert_one(worker_doc.copy())
    await audit(u, "create", "worker", worker_id, None, {"name": body.name})
    return clean(worker_doc)


@api.put("/workers/{worker_id}")
async def update_worker(worker_id: str, body: WorkerUpdate, u=Depends(require_role(Role.admin))):
    old = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Worker not found")
    upd = {k: v for k, v in body.dict().items() if v is not None}
    if upd:
        await db.workers.update_one({"id": worker_id}, {"$set": upd})
        if "status" in upd:
            disabled = upd["status"] != "Active"
            await db.users.update_one({"worker_id": worker_id}, {"$set": {"disabled": disabled}})
        if "name" in upd:
            await db.users.update_one({"worker_id": worker_id}, {"$set": {"name": upd["name"]}})
    await audit(u, "update", "worker", worker_id, old, upd)
    return {**old, **upd}


@api.post("/workers/{worker_id}/reset-password")
async def reset_worker_pwd(worker_id: str, body: dict, u=Depends(require_role(Role.admin))):
    new_pwd = body.get("new_password", "ChangeMe123")
    if len(new_pwd) < 8:
        raise HTTPException(400, "Password too short (min 8 characters)")
    r = await db.users.update_one(
        {"worker_id": worker_id},
        {"$set": {"password_hash": pwd_ctx.hash(new_pwd)}, "$inc": {"token_version": 1}}
    )
    if not r.matched_count:
        raise HTTPException(404, "Worker user not found")
    await audit(u, "reset-password", "worker", worker_id)
    return {"ok": True}


# --- Attendance ----------------------------------------------------------
@api.post("/attendance/check-in")
async def check_in(body: CheckInIn, u=Depends(require_role(Role.worker, Role.admin))):
    worker_id = u.get("worker_id")
    if not worker_id:
        raise HTTPException(400, "Only workers can check in")
    # idempotency — scoped to this worker so a replayed/guessed client_sync_id
    # can't be used to read back another worker's check-in data
    if body.client_sync_id:
        existing = await db.attendance_sessions.find_one(
            {"client_sync_id": body.client_sync_id, "worker_id": worker_id}, {"_id": 0}
        )
        if existing:
            return existing
    # active session guard
    active = await db.attendance_sessions.find_one(
        {"worker_id": worker_id, "check_out_time": None}, {"_id": 0}
    )
    if active:
        raise HTTPException(400, "You have an active site. Check out first.")
    site = await db.sites.find_one({"id": body.site_id}, {"_id": 0})
    if not site or site.get("status") != "Active":
        raise HTTPException(400, "Site is not active")
    assigned = site.get("assigned_worker_ids") or []
    if assigned and worker_id not in assigned:
        raise HTTPException(403, "You are not assigned to this site.")
    geo_distance_m = None
    geo_warning = False
    if site.get("latitude") is not None and site.get("longitude") is not None:
        geo_distance_m = round(haversine_meters(body.latitude, body.longitude, site["latitude"], site["longitude"]))
        geo_warning = geo_distance_m > (site.get("geofence_radius_m") or GEOFENCE_WARNING_METERS)
    session_id = str(uuid.uuid4())
    check_in_time = now_utc().isoformat()
    settings = await db.settings.find_one({"id": "global"}, {"_id": 0}) or {}
    late_flag = False
    late_by_minutes = None
    if settings.get("shift_start_time"):
        grace = settings.get("late_grace_minutes") or 0
        checkin_mins = minutes_since_midnight_ist(check_in_time)
        shift_start_mins = parse_hm_to_minutes(settings["shift_start_time"])
        if checkin_mins > shift_start_mins + grace:
            late_flag = True
            late_by_minutes = checkin_mins - shift_start_mins
    doc = {
        "id": session_id,
        "worker_id": worker_id,
        "worker_name": u.get("name"),
        "site_id": body.site_id,
        "site_name": site["site_name"],
        "check_in_time": check_in_time,
        "check_in_latitude": body.latitude,
        "check_in_longitude": body.longitude,
        "check_in_accuracy": body.accuracy,
        "check_in_geo_warning": geo_warning,
        "check_in_geo_distance_m": geo_distance_m,
        "late_flag": late_flag,
        "late_by_minutes": late_by_minutes,
        "check_out_time": None,
        "check_out_latitude": None,
        "check_out_longitude": None,
        "check_out_geo_warning": False,
        "check_out_geo_distance_m": None,
        "early_leave_flag": False,
        "early_leave_by_minutes": None,
        "work_completed": None,
        "remarks": "",
        "created_at": now_utc().isoformat(),
    }
    if body.client_sync_id:
        doc["client_sync_id"] = body.client_sync_id
    await db.attendance_sessions.insert_one(doc.copy())
    # notify admins — never blocks the check-in itself, just flags it for a human look
    msg = f"{u.get('name')} checked in at {site['site_name']}."
    if geo_warning:
        msg += f" ⚠️ {geo_distance_m}m from the site's recorded location — please verify."
    if late_flag:
        msg += f" ⚠️ {late_by_minutes} min late."
    async for admin in db.users.find({"role": Role.admin.value}, {"_id": 0, "id": 1}):
        await add_notification(admin["id"], "Worker Check-In", msg)
    return clean(doc)


@api.post("/attendance/check-out")
async def check_out(body: CheckOutIn, u=Depends(require_role(Role.worker, Role.admin))):
    sess = await db.attendance_sessions.find_one({"id": body.session_id}, {"_id": 0})
    if not sess:
        raise HTTPException(404, "Session not found")
    if sess["worker_id"] != u.get("worker_id"):
        raise HTTPException(403, "Not your session")
    if sess.get("check_out_time"):
        raise HTTPException(400, "Already checked out")
    site = await db.sites.find_one({"id": sess["site_id"]}, {"_id": 0})
    geo_distance_m = None
    geo_warning = False
    if site and site.get("latitude") is not None and site.get("longitude") is not None:
        geo_distance_m = round(haversine_meters(body.latitude, body.longitude, site["latitude"], site["longitude"]))
        geo_warning = geo_distance_m > ((site or {}).get("geofence_radius_m") or GEOFENCE_WARNING_METERS)
    check_out_time = now_utc().isoformat()
    settings = await db.settings.find_one({"id": "global"}, {"_id": 0}) or {}
    early_leave_flag = False
    early_leave_by_minutes = None
    if settings.get("shift_end_time"):
        grace = settings.get("early_leave_grace_minutes") or 0
        checkout_mins = minutes_since_midnight_ist(check_out_time)
        shift_end_mins = parse_hm_to_minutes(settings["shift_end_time"])
        if checkout_mins < shift_end_mins - grace:
            early_leave_flag = True
            early_leave_by_minutes = shift_end_mins - checkout_mins
    upd = {
        "check_out_time": check_out_time,
        "check_out_latitude": body.latitude,
        "check_out_longitude": body.longitude,
        "check_out_accuracy": body.accuracy,
        "check_out_geo_warning": geo_warning,
        "check_out_geo_distance_m": geo_distance_m,
        "early_leave_flag": early_leave_flag,
        "early_leave_by_minutes": early_leave_by_minutes,
        "work_completed": body.work_completed,
        "remarks": body.remarks or "",
    }
    await db.attendance_sessions.update_one({"id": body.session_id}, {"$set": upd})
    msg = f"{u.get('name')} checked out from {sess['site_name']}."
    if geo_warning:
        msg += f" ⚠️ {geo_distance_m}m from the site's recorded location — please verify."
    if early_leave_flag:
        msg += f" ⚠️ Left {early_leave_by_minutes} min early."
    async for admin in db.users.find({"role": Role.admin.value}, {"_id": 0, "id": 1}):
        await add_notification(admin["id"], "Worker Check-Out", msg)
    return {**sess, **upd}


class ForceCheckoutIn(BaseModel):
    remarks: Optional[str] = ""


@api.post("/attendance/{session_id}/force-checkout")
async def force_checkout(session_id: str, body: ForceCheckoutIn, u=Depends(require_role(Role.admin))):
    """Admin-initiated close for a session the worker forgot to check out of —
    no GPS captured since the admin wasn't physically there."""
    sess = await db.attendance_sessions.find_one({"id": session_id}, {"_id": 0})
    if not sess:
        raise HTTPException(404, "Session not found")
    if sess.get("check_out_time"):
        raise HTTPException(400, "Already checked out")
    upd = {
        "check_out_time": now_utc().isoformat(),
        "work_completed": None,  # unobserved — admin wasn't there to confirm
        "remarks": body.remarks or "Closed by admin — worker did not check out.",
        "force_closed_by": u["id"],
        "force_closed_by_name": u.get("name") or u.get("username"),
    }
    await db.attendance_sessions.update_one({"id": session_id}, {"$set": upd})
    await audit(u, "force-checkout", "attendance", session_id, sess, upd)
    wu = await db.users.find_one({"worker_id": sess["worker_id"]}, {"_id": 0, "id": 1})
    if wu:
        await add_notification(
            wu["id"], "Checked Out By Admin",
            f"Your session at {sess.get('site_name')} was closed by an admin — you didn't check out.",
        )
    return {**sess, **upd}


@api.get("/attendance")
async def list_attendance(
    u=Depends(get_user),
    worker_id: Optional[str] = None,
    site_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    q = {}
    if u["role"] == Role.worker.value:
        q["worker_id"] = u.get("worker_id")
    elif worker_id:
        q["worker_id"] = worker_id
    if site_id:
        q["site_id"] = site_id
    if date_from:
        q.setdefault("check_in_time", {})["$gte"] = date_from
    if date_to:
        q.setdefault("check_in_time", {})["$lte"] = date_to + "T23:59:59"
    items = await db.attendance_sessions.find(q, {"_id": 0}).sort("check_in_time", -1).to_list(1000)
    return items


@api.get("/attendance/active")
async def active_session(u=Depends(require_role(Role.worker, Role.admin))):
    if not u.get("worker_id"):
        return None
    s = await db.attendance_sessions.find_one(
        {"worker_id": u["worker_id"], "check_out_time": None}, {"_id": 0}
    )
    return s


# --- Work reports --------------------------------------------------------
async def _initial_excel_sync_status() -> str:
    cfg = await db.excel_integration.find_one({"id": "global"}, {"_id": 0, "enabled": 1})
    return ExcelSyncStatus.pending.value if (cfg and cfg.get("enabled")) else ExcelSyncStatus.not_applicable.value


async def _next_entry_id() -> str:
    """The app-wide idempotency key used as the Excel sync key (master prompt's
    'Entry ID' requirement) — independent of Excel's own Work ID/Entry ID, which
    are minted separately, in Excel, at sync time."""
    r = await db.counters.find_one_and_update(
        {"id": "work-report-entry"}, {"$inc": {"value": 1}}, upsert=True, return_document=True,
    )
    return f"ENTRY-{(r or {}).get('value', 1):06d}"


async def _activate_worker_automation(worker_id: str) -> None:
    """A worker's automation starts on their own first VALID app work entry —
    never on login, never on a fixed calendar date, and never shared across
    workers. 'Valid' means it passed submission validation (this function only
    runs after a work_reports insert has already succeeded) — it is
    deliberately NOT gated on later admin approval, since approval is a
    business decision about the work, not evidence of whether the worker
    started using the app."""
    worker = await db.workers.find_one({"id": worker_id}, {"_id": 0, "automation_enabled": 1})
    if worker and worker.get("automation_enabled"):
        return  # already activated — never move the date on a later entry
    ts = now_utc().isoformat()
    await db.workers.update_one(
        {"id": worker_id, "automation_enabled": {"$ne": True}},
        {"$set": {"automation_enabled": True, "first_valid_entry_at": ts, "automation_started_at": ts}},
    )


@api.post("/work-reports")
async def create_report(body: WorkReportIn, u=Depends(require_role(Role.worker, Role.admin))):
    # idempotency — scoped to this worker, same reasoning as check_in above
    if body.client_sync_id:
        existing = await db.work_reports.find_one(
            {"client_sync_id": body.client_sync_id, "worker_id": u.get("worker_id")}, {"_id": 0}
        )
        if existing:
            return existing
    sess = await db.attendance_sessions.find_one({"id": body.session_id}, {"_id": 0})
    if not sess:
        raise HTTPException(404, "Session not found")
    if sess["worker_id"] != u.get("worker_id"):
        raise HTTPException(403, "Not your session")
    if len(body.before_photos) < 1 or len(body.after_photos) < 1:
        raise HTTPException(400, "At least 1 before and 1 after photo required")
    if any(_photo_too_large(p) for p in body.before_photos + body.after_photos):
        raise HTTPException(400, "One or more photos exceed the size limit (~9MB each)")
    for i, p in enumerate(body.before_photos):
        _validate_photo(p, f"Before photo {i + 1}")
    for i, p in enumerate(body.after_photos):
        _validate_photo(p, f"After photo {i + 1}")
    # Enforce each door type's assigned quantity at this site, if the admin set one.
    # Checked before touching photos so a worker gets fast feedback instead of a
    # rejection after already uploading evidence for a count that was never going in.
    site = await db.sites.find_one({"id": sess["site_id"]}, {"_id": 0})
    targets = (site or {}).get("target_doors", {}) or {}
    for item in body.doors:
        target = targets.get(item.door_type_id)
        if target is None:
            continue
        installed = await get_installed_count(sess["site_id"], item.door_type_id)
        remaining = target - installed
        if item.count > remaining:
            door_type = await db.door_types.find_one({"id": item.door_type_id}, {"_id": 0})
            type_name = door_type["name"] if door_type else "this door type"
            raise HTTPException(
                400,
                f"Only {max(remaining, 0)} {type_name} remaining at this site "
                f"(target: {target}, already recorded: {installed}). Ask an admin to "
                f"raise the target if this is intentional.",
            )
    # server-computed breakdown & total (never trust client-supplied rates)
    door_breakdown = []
    door_count = 0
    total = 0.0
    for item in body.doors:
        name, rate = await get_door_type_rate(item.door_type_id, sess["site_id"])
        subtotal = round(item.count * rate, 2)
        door_breakdown.append({
            "door_type_id": item.door_type_id,
            "door_type_name": name,
            "rate": rate,
            "count": item.count,
            "subtotal": subtotal,
        })
        door_count += item.count
        total += subtotal
    total = round(total, 2)
    # Watermark all photos server-side, then store each as its own report_photos
    # document rather than embedding them in the report — a worker submitting
    # 15-20 photos per side would otherwise risk the report document itself
    # bumping into MongoDB's 16MB per-document ceiling.
    report_id = str(uuid.uuid4())
    lat_ci = sess.get("check_in_latitude")
    lng_ci = sess.get("check_in_longitude")

    async def _store_photos(photos: list, kind: str) -> list:
        ids = []
        for p in photos:
            watermarked = _watermark_photo(p, sess.get("worker_name") or "-", sess.get("site_name") or "-", lat_ci, lng_ci)
            photo_id = str(uuid.uuid4())
            await db.report_photos.insert_one({
                "id": photo_id, "report_id": report_id, "kind": kind,
                "data": watermarked, "created_at": now_utc().isoformat(),
            })
            ids.append(photo_id)
        return ids

    before_photo_ids = await _store_photos(body.before_photos, "before")
    after_photo_ids = await _store_photos(body.after_photos, "after")
    entry_id = await _next_entry_id()
    doc = {
        "id": report_id,
        "entry_id": entry_id,
        "session_id": body.session_id,
        "worker_id": sess["worker_id"],
        "worker_name": sess.get("worker_name"),
        "site_id": sess["site_id"],
        "site_name": sess.get("site_name"),
        "work_date": now_utc().date().isoformat(),
        "door_count": door_count,
        "rate_per_door": round(total / door_count, 2) if door_count else 0,
        "door_breakdown": door_breakdown,
        "total_amount": total,
        "notes": body.notes or "",
        "work_completed": body.work_completed,
        "before_photo_ids": before_photo_ids,
        "after_photo_ids": after_photo_ids,
        "materials": body.materials or [],
        "approval_status": ApprovalStatus.pending.value,
        "approval_remarks": "",
        "approved_by": None,
        "approved_at": None,
        "created_at": now_utc().isoformat(),
        # Excel Integration — nothing syncs until the report is Approved (Excel's
        # T_WORK has no "pending review" concept; see the audit report). The DB
        # stays the source of truth regardless of sync outcome.
        "excel_sync_status": ExcelSyncStatus.not_applicable.value,
        "excel_sync_attempts": 0,
        "excel_last_synced_at": None,
        "excel_last_error": None,
        "excel_work_ids": [],
    }
    if body.client_sync_id:
        doc["client_sync_id"] = body.client_sync_id
    await db.work_reports.insert_one(doc.copy())
    await _activate_worker_automation(sess["worker_id"])
    async for admin in db.users.find({"role": Role.admin.value}, {"_id": 0, "id": 1}):
        await add_notification(admin["id"], "New Work Report", f"{sess.get('worker_name')} submitted a report ({door_count} doors).")
    return clean(doc)


@api.put("/work-reports/{report_id}")
async def update_report(report_id: str, body: WorkReportUpdateIn, u=Depends(require_role(Role.worker))):
    """Worker self-edit of their own still-open report — fixes the 'wait for an
    admin to notice a typo' gap. Only while Pending or Correction; editing a
    Correction resets it to Pending since it needs a fresh look either way."""
    r = await db.work_reports.find_one({"id": report_id}, {"_id": 0})
    if not r:
        raise HTTPException(404, "Not found")
    if r["worker_id"] != u.get("worker_id"):
        raise HTTPException(403, "Not your report")
    if r["approval_status"] not in (ApprovalStatus.pending.value, ApprovalStatus.correction.value):
        raise HTTPException(400, "Only a Pending or Correction report can be edited")
    site_id = r["site_id"]
    site = await db.sites.find_one({"id": site_id}, {"_id": 0})
    targets = (site or {}).get("target_doors", {}) or {}
    for item in body.doors:
        target = targets.get(item.door_type_id)
        if target is None:
            continue
        installed = await get_installed_count(site_id, item.door_type_id, exclude_report_id=report_id)
        remaining = target - installed
        if item.count > remaining:
            door_type = await db.door_types.find_one({"id": item.door_type_id}, {"_id": 0})
            type_name = door_type["name"] if door_type else "this door type"
            raise HTTPException(
                400,
                f"Only {max(remaining, 0)} {type_name} remaining at this site "
                f"(target: {target}, already recorded: {installed}).",
            )
    door_breakdown = []
    door_count = 0
    total = 0.0
    for item in body.doors:
        name, rate = await get_door_type_rate(item.door_type_id, site_id)
        subtotal = round(item.count * rate, 2)
        door_breakdown.append({
            "door_type_id": item.door_type_id,
            "door_type_name": name,
            "rate": rate,
            "count": item.count,
            "subtotal": subtotal,
        })
        door_count += item.count
        total += subtotal
    total = round(total, 2)
    upd = {
        "door_count": door_count,
        "rate_per_door": round(total / door_count, 2) if door_count else 0,
        "door_breakdown": door_breakdown,
        "total_amount": total,
        "notes": body.notes or "",
        "work_completed": body.work_completed,
        "approval_status": ApprovalStatus.pending.value,
    }
    await db.work_reports.update_one({"id": report_id}, {"$set": upd})
    await audit(u, "update", "work_report", report_id, r, upd)
    async for admin in db.users.find({"role": Role.admin.value}, {"_id": 0, "id": 1}):
        await add_notification(admin["id"], "Work Report Updated", f"{r.get('worker_name')} updated a report ({door_count} doors).")
    return {**r, **upd}


@api.delete("/work-reports/{report_id}")
async def withdraw_report(report_id: str, u=Depends(require_role(Role.worker))):
    """Worker withdraws their own report entirely, e.g. to redo it with different
    photos — self-edit above deliberately can't touch photos, this is the escape
    hatch for when photos are the thing that's wrong."""
    r = await db.work_reports.find_one({"id": report_id}, {"_id": 0})
    if not r:
        raise HTTPException(404, "Not found")
    if r["worker_id"] != u.get("worker_id"):
        raise HTTPException(403, "Not your report")
    if r["approval_status"] not in (ApprovalStatus.pending.value, ApprovalStatus.correction.value):
        raise HTTPException(400, "Only a Pending or Correction report can be withdrawn")
    await db.work_reports.delete_one({"id": report_id})
    await db.report_photos.delete_many({"report_id": report_id})
    await audit(u, "delete", "work_report", report_id, r, None)
    return {"ok": True}


@api.get("/work-reports/{report_id}/photos")
async def get_report_photos(report_id: str, u=Depends(get_user)):
    r = await db.work_reports.find_one({"id": report_id}, {"_id": 0})
    if not r:
        raise HTTPException(404, "Not found")
    if u["role"] == Role.worker.value and r["worker_id"] != u.get("worker_id"):
        raise HTTPException(403, "Not your report")
    photos = await db.report_photos.find({"report_id": report_id}, {"_id": 0}).sort("created_at", 1).to_list(200)
    if photos:
        return {
            "before": [p["data"] for p in photos if p["kind"] == "before"],
            "after": [p["data"] for p in photos if p["kind"] == "after"],
        }
    # Legacy reports created before photos moved into their own collection still
    # have them embedded directly on the report document — fall back to that.
    return {"before": r.get("before_photos", []), "after": r.get("after_photos", [])}


@api.get("/work-reports")
async def list_reports(
    u=Depends(get_user),
    approval_status: Optional[str] = None,
    worker_id: Optional[str] = None,
    site_id: Optional[str] = None,
):
    q = {}
    if u["role"] == Role.worker.value:
        q["worker_id"] = u.get("worker_id")
    elif worker_id:
        q["worker_id"] = worker_id
    if approval_status:
        q["approval_status"] = approval_status
    if site_id:
        q["site_id"] = site_id
    items = await db.work_reports.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items


@api.get("/work-reports/{report_id}")
async def get_report(report_id: str, u=Depends(get_user)):
    r = await db.work_reports.find_one({"id": report_id}, {"_id": 0})
    if not r:
        raise HTTPException(404, "Not found")
    if u["role"] == Role.worker.value and r["worker_id"] != u.get("worker_id"):
        raise HTTPException(403, "Forbidden")
    return r


async def _approval_action(report_id: str, new_status: str, body: ApprovalIn, u: dict):
    r = await db.work_reports.find_one({"id": report_id}, {"_id": 0})
    if not r:
        raise HTTPException(404, "Not found")
    upd = {
        "approval_status": new_status,
        "approval_remarks": body.remarks or "",
        "approved_by": u["id"],
        "approved_by_name": u.get("name") or u.get("username"),
        "approved_at": now_utc().isoformat(),
    }
    if new_status == ApprovalStatus.approved.value:
        # allow admin edit at approval time
        dc = body.edited_door_count if body.edited_door_count is not None else r["door_count"]
        rate = body.edited_rate if body.edited_rate is not None else r["rate_per_door"]
        upd["door_count"] = dc
        upd["rate_per_door"] = rate
        upd["total_amount"] = round(dc * rate, 2)
        # Excel's T_WORK has no "pending review" concept (verified in the audit) —
        # a report only ever syncs once it's Approved. Re-approving after an edit
        # or an unapprove/reapprove cycle re-syncs the SAME row(s) via excel_work_ids,
        # it never creates a duplicate.
        upd["excel_sync_status"] = await _initial_excel_sync_status()
    elif new_status == ApprovalStatus.rejected.value and r.get("excel_work_ids"):
        # Already-synced work got rejected after the fact — the row must be
        # flagged, never silently deleted from Excel.
        upd["excel_sync_status"] = await _initial_excel_sync_status()
    await db.work_reports.update_one({"id": report_id}, {"$set": upd})
    await audit(u, f"report-{new_status.lower()}", "work_report", report_id, r, upd)
    # notify worker
    worker_user = await db.users.find_one({"worker_id": r["worker_id"]}, {"_id": 0, "id": 1})
    if worker_user:
        title_map = {
            ApprovalStatus.approved.value: "Work Approved",
            ApprovalStatus.rejected.value: "Work Rejected",
            ApprovalStatus.correction.value: "Correction Requested",
        }
        msg_map = {
            ApprovalStatus.approved.value: f"Your work report has been approved. Amount: ₹{upd.get('total_amount', r['total_amount'])}.",
            ApprovalStatus.rejected.value: f"Your work report was rejected. {body.remarks or ''}",
            ApprovalStatus.correction.value: f"Correction requested. {body.remarks or ''}",
        }
        await add_notification(worker_user["id"], title_map[new_status], msg_map[new_status])
    return {**r, **upd}


@api.post("/work-reports/{report_id}/approve")
async def approve_report(report_id: str, body: ApprovalIn, u=Depends(require_role(Role.admin))):
    return await _approval_action(report_id, ApprovalStatus.approved.value, body, u)


@api.post("/work-reports/{report_id}/reject")
async def reject_report(report_id: str, body: ApprovalIn, u=Depends(require_role(Role.admin))):
    return await _approval_action(report_id, ApprovalStatus.rejected.value, body, u)


@api.post("/work-reports/{report_id}/correction")
async def correction_report(report_id: str, body: ApprovalIn, u=Depends(require_role(Role.admin))):
    return await _approval_action(report_id, ApprovalStatus.correction.value, body, u)


@api.post("/work-reports/{report_id}/unapprove")
async def unapprove_report(report_id: str, body: ApprovalIn, u=Depends(require_role(Role.admin))):
    """Reopen a mistakenly-approved report back to Pending. Restores door_count/
    total_amount to what was actually submitted (from door_breakdown), undoing any
    edit the admin made at approval time — reopening should mean starting review
    over, not keeping half of a correction. Does not touch any payment already
    recorded against it; that reconciliation is a separate manual step."""
    r = await db.work_reports.find_one({"id": report_id}, {"_id": 0})
    if not r:
        raise HTTPException(404, "Not found")
    if r.get("approval_status") != ApprovalStatus.approved.value:
        raise HTTPException(400, "Only an Approved report can be reopened")
    breakdown = r.get("door_breakdown") or []
    upd = {
        "approval_status": ApprovalStatus.pending.value,
        "approval_remarks": body.remarks or "Reopened by admin for re-review",
        "approved_by": None,
        "approved_by_name": None,
        "approved_at": None,
    }
    if breakdown:
        orig_count = sum(item.get("count", 0) for item in breakdown)
        orig_total = round(sum(item.get("subtotal", 0) for item in breakdown), 2)
        upd["door_count"] = orig_count
        upd["total_amount"] = orig_total
        upd["rate_per_door"] = round(orig_total / orig_count, 2) if orig_count else 0
    await db.work_reports.update_one({"id": report_id}, {"$set": upd})
    await audit(u, "report-reopened", "work_report", report_id, r, upd)
    wu = await db.users.find_one({"worker_id": r["worker_id"]}, {"_id": 0, "id": 1})
    if wu:
        await add_notification(
            wu["id"], "Report Reopened",
            f"Your approved report was reopened for review. {body.remarks or ''}",
        )
    return {**r, **upd}


# --- Payments ------------------------------------------------------------
DUPLICATE_PAYMENT_WINDOW_MINUTES = 30  # catches stale-data re-runs, distinct from client_sync_id's exact-retry case


async def find_recent_duplicate_payment(worker_id: str, amount: float) -> Optional[dict]:
    """A same-worker, same-amount payment recorded moments ago is almost always
    a duplicate (stale pending-amount data, a re-run of the payout screen) —
    never blocked, just flagged so the admin can double-check."""
    cutoff = (now_utc() - timedelta(minutes=DUPLICATE_PAYMENT_WINDOW_MINUTES)).isoformat()
    return await db.payments.find_one(
        {"worker_id": worker_id, "amount": round(amount, 2), "created_at": {"$gt": cutoff}},
        {"_id": 0},
    )


@api.post("/payments")
async def create_payment(body: PaymentIn, u=Depends(require_role(Role.admin))):
    if body.amount <= 0:
        raise HTTPException(400, "Amount must be positive")
    # idempotency — a double-click or slow-network retry must not double-pay a worker
    if body.client_sync_id:
        existing = await db.payments.find_one({"client_sync_id": body.client_sync_id}, {"_id": 0})
        if existing:
            return existing
    worker = await db.workers.find_one({"id": body.worker_id}, {"_id": 0})
    if not worker:
        raise HTTPException(404, "Worker not found")
    duplicate = await find_recent_duplicate_payment(body.worker_id, body.amount)
    doc = {
        "id": str(uuid.uuid4()),
        "worker_id": body.worker_id,
        "worker_name": worker["name"],
        "payment_date": body.payment_date or now_utc().date().isoformat(),
        "amount": round(body.amount, 2),
        "payment_method": body.payment_method,
        "transaction_reference": body.transaction_reference or "",
        "notes": body.notes or "",
        "created_by": u["id"],
        "created_at": now_utc().isoformat(),
        "excel_sync_status": await _initial_excel_sync_status(),
        "excel_sync_attempts": 0,
        "excel_last_synced_at": None,
        "excel_last_error": None,
        "excel_entry_id": None,
    }
    if body.client_sync_id:
        doc["client_sync_id"] = body.client_sync_id
    await db.payments.insert_one(doc.copy())
    await audit(u, "create", "payment", doc["id"], None, doc)
    # notify worker
    worker_user = await db.users.find_one({"worker_id": body.worker_id}, {"_id": 0, "id": 1})
    if worker_user:
        await add_notification(worker_user["id"], "Payment Added", f"Payment of ₹{doc['amount']} recorded.")
    result = clean(doc)
    if duplicate:
        result["duplicate_warning"] = (
            f"Another ₹{duplicate['amount']} payment to {worker['name']} was recorded "
            f"at {formatted_ist(duplicate['created_at'])}. Please double-check this isn't a repeat."
        )
    return result


@api.get("/payments")
async def list_payments(u=Depends(get_user), worker_id: Optional[str] = None):
    q = {}
    if u["role"] == Role.worker.value:
        q["worker_id"] = u.get("worker_id")
    elif worker_id:
        q["worker_id"] = worker_id
    items = await db.payments.find(q, {"_id": 0}).sort("payment_date", -1).to_list(500)
    return items


# --- Notifications -------------------------------------------------------
@api.get("/notifications")
async def list_notifications(u=Depends(get_user)):
    items = await db.notifications.find({"user_id": u["id"]}, {"_id": 0}).sort("created_at", -1).limit(100).to_list(100)
    return items


@api.post("/notifications/{nid}/read")
async def read_notif(nid: str, u=Depends(get_user)):
    await db.notifications.update_one({"id": nid, "user_id": u["id"]}, {"$set": {"is_read": True}})
    return {"ok": True}


@api.post("/notifications/read-all")
async def read_all(u=Depends(get_user)):
    await db.notifications.update_many({"user_id": u["id"]}, {"$set": {"is_read": True}})
    return {"ok": True}


# --- Dashboards / Ledger -------------------------------------------------
async def _worker_earnings(worker_id: str):
    approved = await db.work_reports.aggregate([
        {"$match": {"worker_id": worker_id, "approval_status": ApprovalStatus.approved.value}},
        {"$group": {"_id": None, "total": {"$sum": "$total_amount"}, "doors": {"$sum": "$door_count"}}}
    ]).to_list(1)
    paid = await db.payments.aggregate([
        {"$match": {"worker_id": worker_id}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    approved_amt = approved[0]["total"] if approved else 0
    doors = approved[0]["doors"] if approved else 0
    paid_amt = paid[0]["total"] if paid else 0
    return {
        "approved_earnings": round(approved_amt, 2),
        "paid_amount": round(paid_amt, 2),
        "pending_amount": round(approved_amt - paid_amt, 2),
        "total_doors": doors,
    }


@api.get("/worker/dashboard")
async def worker_dashboard(u=Depends(require_role(Role.worker, Role.admin))):
    if not u.get("worker_id"):
        raise HTTPException(400, "Not a worker")
    wid = u["worker_id"]
    today = now_utc().date().isoformat()
    month_start = now_utc().date().replace(day=1).isoformat()
    active = await db.attendance_sessions.find_one({"worker_id": wid, "check_out_time": None}, {"_id": 0})
    today_sessions = await db.attendance_sessions.count_documents({
        "worker_id": wid, "check_in_time": {"$gte": today}
    })
    today_reports = await db.work_reports.aggregate([
        {"$match": {"worker_id": wid, "work_date": today}},
        {"$group": {"_id": None, "doors": {"$sum": "$door_count"},
                    "approved_amount": {"$sum": {"$cond": [{"$eq": ["$approval_status", "Approved"]}, "$total_amount", 0]}}}}
    ]).to_list(1)
    month_reports = await db.work_reports.aggregate([
        {"$match": {"worker_id": wid, "work_date": {"$gte": month_start}}},
        {"$group": {"_id": None, "doors": {"$sum": "$door_count"},
                    "approved_amount": {"$sum": {"$cond": [{"$eq": ["$approval_status", "Approved"]}, "$total_amount", 0]}}}}
    ]).to_list(1)
    month_sites = await db.attendance_sessions.distinct("site_id", {
        "worker_id": wid, "check_in_time": {"$gte": month_start}
    })
    earnings = await _worker_earnings(wid)
    return {
        "check_in_status": "checked_in" if active else "checked_out",
        "active_session": active,
        "today": {
            "sites_visited": today_sessions,
            "doors_installed": today_reports[0]["doors"] if today_reports else 0,
            "approved_amount": round(today_reports[0]["approved_amount"], 2) if today_reports else 0,
        },
        "month": {
            "sites_visited": len(month_sites),
            "doors_installed": month_reports[0]["doors"] if month_reports else 0,
            "approved_amount": round(month_reports[0]["approved_amount"], 2) if month_reports else 0,
            **earnings,
        },
    }


@api.get("/admin/dashboard")
async def admin_dashboard(u=Depends(require_role(Role.admin))):
    today = now_utc().date().isoformat()
    month_start = now_utc().date().replace(day=1).isoformat()
    year_start = now_utc().date().replace(month=1, day=1).isoformat()
    total_workers = await db.workers.count_documents({})
    active_sites = await db.sites.count_documents({"status": "Active"})
    active_now = await db.attendance_sessions.count_documents({"check_out_time": None})
    checked_in_today_ids = await db.attendance_sessions.distinct("worker_id", {"check_in_time": {"$gte": today}})
    sites_today = await db.attendance_sessions.distinct("site_id", {"check_in_time": {"$gte": today}})

    async def doors_between(start):
        r = await db.work_reports.aggregate([
            {"$match": {"work_date": {"$gte": start}}},
            {"$group": {"_id": None, "doors": {"$sum": "$door_count"}}}
        ]).to_list(1)
        return r[0]["doors"] if r else 0

    approved_agg = await db.work_reports.aggregate([
        {"$match": {"approval_status": "Approved"}},
        {"$group": {"_id": None, "total": {"$sum": "$total_amount"}}}
    ]).to_list(1)
    paid_agg = await db.payments.aggregate([
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    approved_total = approved_agg[0]["total"] if approved_agg else 0
    paid_total = paid_agg[0]["total"] if paid_agg else 0
    pending_reports = await db.work_reports.count_documents({"approval_status": "Pending"})

    return {
        "workers": {
            "total": total_workers,
            "active_today": len(checked_in_today_ids),
            "currently_checked_in": active_now,
            "inactive_today": total_workers - len(checked_in_today_ids),
        },
        "sites": {
            "total_active": active_sites,
            "with_workers_today": len(sites_today),
        },
        "doors": {
            "today": await doors_between(today),
            "month": await doors_between(month_start),
            "year": await doors_between(year_start),
        },
        "financial": {
            "approved_total": round(approved_total, 2),
            "paid_total": round(paid_total, 2),
            "pending_total": round(approved_total - paid_total, 2),
        },
        "pending_reports": pending_reports,
    }


@api.get("/workers/{worker_id}/ledger")
async def worker_ledger(worker_id: str, u=Depends(require_role(Role.admin, Role.worker))):
    if u["role"] == Role.worker.value and u.get("worker_id") != worker_id:
        raise HTTPException(403, "Forbidden")
    entries = []
    async for r in db.work_reports.find(
        {"worker_id": worker_id, "approval_status": "Approved"}, {"_id": 0}
    ):
        entries.append({
            "date": r["work_date"],
            "description": f"{r['door_count']} doors @ ₹{r['rate_per_door']}",
            "site": r.get("site_name"),
            "type": "credit",
            "amount": r["total_amount"],
        })
    async for p in db.payments.find({"worker_id": worker_id}, {"_id": 0}):
        entries.append({
            "date": p["payment_date"],
            "description": f"Payment via {p['payment_method']}" + (f" ({p['transaction_reference']})" if p.get("transaction_reference") else ""),
            "site": None,
            "type": "debit",
            "amount": p["amount"],
        })
    entries.sort(key=lambda x: x["date"])
    balance = 0
    for e in entries:
        if e["type"] == "credit":
            balance += e["amount"]
        else:
            balance -= e["amount"]
        e["balance"] = round(balance, 2)
    earnings = await _worker_earnings(worker_id)
    return {"entries": entries, "summary": earnings}


@api.get("/workers/{worker_id}/summary")
async def worker_summary(worker_id: str, u=Depends(require_role(Role.admin, Role.worker))):
    if u["role"] == Role.worker.value and u.get("worker_id") != worker_id:
        raise HTTPException(403, "Forbidden")
    return await _worker_earnings(worker_id)


@api.get("/workers/{worker_id}/performance")
async def worker_performance(worker_id: str, u=Depends(require_role(Role.admin, Role.worker))):
    """Lifetime performance metrics — only ones honestly computable from what we
    actually track. No fabricated "attendance %"/score: that would need a concept
    of an expected daily roster this system doesn't have."""
    if u["role"] == Role.worker.value and u.get("worker_id") != worker_id:
        raise HTTPException(403, "Forbidden")
    worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    if not worker:
        raise HTTPException(404, "Worker not found")

    total_doors = 0
    total_earned = 0.0
    working_days = set()
    approved_count = 0
    rejected_count = 0
    pending_count = 0
    async for r in db.work_reports.find({"worker_id": worker_id}, {"_id": 0}):
        status = r.get("approval_status")
        if status == "Approved":
            approved_count += 1
            total_doors += r.get("door_count", 0)
            total_earned += float(r.get("total_amount", 0))
            working_days.add(r.get("work_date"))
        elif status == "Rejected":
            rejected_count += 1
        elif status in ("Pending", "Correction"):
            pending_count += 1

    total_sessions = 0
    late_sessions = 0
    async for a in db.attendance_sessions.find(
        {"worker_id": worker_id, "check_out_time": {"$ne": None}}, {"_id": 0, "late_flag": 1}
    ):
        total_sessions += 1
        if a.get("late_flag"):
            late_sessions += 1

    paid = 0.0
    async for p in db.payments.find({"worker_id": worker_id}, {"_id": 0, "amount": 1}):
        paid += float(p.get("amount", 0))

    return {
        "worker_id": worker_id,
        "worker_name": worker["name"],
        "total_doors": total_doors,
        "working_days": len(working_days),
        "avg_doors_per_day": round(total_doors / len(working_days), 1) if working_days else 0,
        "approved_reports": approved_count,
        "rejected_reports": rejected_count,
        "pending_reports": pending_count,
        "completed_sessions": total_sessions,
        "on_time_pct": round(100 * (total_sessions - late_sessions) / total_sessions, 1) if total_sessions else None,
        "total_earned": round(total_earned, 2),
        "total_paid": round(paid, 2),
        "pending_amount": round(total_earned - paid, 2),
    }


# --- Settings ------------------------------------------------------------
@api.get("/settings")
async def get_settings(u=Depends(get_user)):
    s = await db.settings.find_one({"id": "global"}, {"_id": 0})
    if not s:
        s = {"id": "global", "company_name": "JONAH ENTERPRISES", "currency": "₹", "default_rate": 250.0}
    return s


_HM_RE = re.compile(r"^([01]\d|2[0-3]):[0-5]\d$")


def _validate_shift_time(value: Optional[str], label: str):
    # Blank clears/disables the flag (existing behavior) — only a non-empty,
    # malformed value is rejected. Without this, a bad save here would throw
    # an uncaught error inside every subsequent check-in/check-out.
    if value and not _HM_RE.match(value):
        raise HTTPException(400, f"{label} must be in HH:MM 24-hour format (e.g. 09:00)")


@api.put("/settings")
async def update_settings(body: SettingsIn, u=Depends(require_role(Role.admin))):
    _validate_shift_time(body.shift_start_time, "Shift start")
    _validate_shift_time(body.shift_end_time, "Shift end")
    if body.logo_base64:
        if _photo_too_large(body.logo_base64):
            raise HTTPException(400, "Logo exceeds the size limit (~9MB)")
        _validate_photo(body.logo_base64, "Logo")
    old = await db.settings.find_one({"id": "global"}, {"_id": 0})
    upd = {k: v for k, v in body.dict().items() if v is not None}
    await db.settings.update_one({"id": "global"}, {"$set": upd}, upsert=True)
    s = await db.settings.find_one({"id": "global"}, {"_id": 0})
    await audit(u, "update", "settings", "global", old, upd)
    return s


# --- Push registration ---------------------------------------------------
@api.post("/register-push", status_code=201)
async def register_push(body: RegisterPushIn, u=Depends(get_user)):
    """Upsert the user's device token in the Emergent push service.
    We require an authenticated user + body.user_id must match to prevent spoofing."""
    if body.user_id != u["id"]:
        raise HTTPException(403, "user_id mismatch")
    try:
        resp = await _push_client.post("/api/v1/push/users/register", json=body.dict())
        if resp.status_code == 401:
            logger.warning("EMERGENT_PUSH_KEY missing or invalid on register")
        # We do NOT persist the token locally — SuprSend resolves it internally.
    except Exception as e:
        logger.warning(f"push register non-blocking failure: {e}")
    return {"status": "registered"}


# --- Weekly Payout Preview -----------------------------------------------
@api.get("/admin/payout-preview")
async def payout_preview(u=Depends(require_role(Role.admin))):
    """Returns one row per worker with a positive pending balance so an admin
    can batch-process this week's payouts in a single screen."""
    # aggregate approved per worker
    approved = {r["_id"]: r for r in await db.work_reports.aggregate([
        {"$match": {"approval_status": "Approved"}},
        {"$group": {"_id": "$worker_id", "total": {"$sum": "$total_amount"}}}
    ]).to_list(1000)}
    paid = {r["_id"]: r for r in await db.payments.aggregate([
        {"$group": {"_id": "$worker_id", "total": {"$sum": "$amount"},
                    "last_date": {"$max": "$payment_date"}}}
    ]).to_list(1000)}
    # One query for everyone's recent payment methods instead of one query per
    # worker inside the loop below — $group's $push preserves the order of the
    # preceding $sort, so each worker's array is already most-recent-first.
    methods_agg = await db.payments.aggregate([
        {"$sort": {"payment_date": -1}},
        {"$group": {"_id": "$worker_id", "methods": {"$push": "$payment_method"}}},
    ]).to_list(1000)
    methods_by_worker = {r["_id"]: [m for m in r["methods"][:5] if m] for r in methods_agg}
    workers = await db.workers.find({"status": "Active"}, {"_id": 0}).to_list(1000)
    rows = []
    for w in workers:
        a = float(approved.get(w["id"], {}).get("total", 0) or 0)
        p = float(paid.get(w["id"], {}).get("total", 0) or 0)
        pending = round(a - p, 2)
        if pending <= 0:
            continue
        # preferred method = most-used among the worker's last 5 payments
        methods = methods_by_worker.get(w["id"], [])
        preferred = max(set(methods), key=methods.count) if methods else "UPI"
        rows.append({
            "worker_id": w["id"],
            "worker_name": w["name"],
            "employee_id": w.get("employee_id"),
            "mobile": w.get("mobile"),
            "approved_total": round(a, 2),
            "paid_total": round(p, 2),
            "pending_amount": pending,
            "preferred_method": preferred,
            "last_paid_date": paid.get(w["id"], {}).get("last_date"),
        })
    rows.sort(key=lambda r: r["pending_amount"], reverse=True)
    total = round(sum(r["pending_amount"] for r in rows), 2)
    return {"rows": rows, "total_pending": total, "worker_count": len(rows)}


class BatchPayItem(BaseModel):
    worker_id: str
    amount: float = Field(gt=0)
    payment_method: str = "UPI"
    transaction_reference: Optional[str] = ""
    notes: Optional[str] = ""


class BatchPayIn(BaseModel):
    payments: List[BatchPayItem]
    payment_date: Optional[str] = None
    client_sync_id: Optional[str] = None  # idempotency for the whole batch, distinct from
    # per-payment client_sync_id since one batch legitimately creates several payment docs


@api.post("/payments/batch")
async def batch_payments(body: BatchPayIn, u=Depends(require_role(Role.admin))):
    """Create multiple payment entries in one call — used by the weekly payout screen."""
    if not body.payments:
        raise HTTPException(400, "No payments provided")
    # idempotency — a double-click on "Pay Selected" must not re-run the whole batch
    if body.client_sync_id:
        existing = await db.payments.find({"batch_sync_id": body.client_sync_id}, {"_id": 0}).to_list(1000)
        if existing:
            return {"created": len(existing), "payments": existing}
    date = body.payment_date or now_utc().date().isoformat()
    created = []
    warnings = []
    for item in body.payments:
        worker = await db.workers.find_one({"id": item.worker_id}, {"_id": 0})
        if not worker:
            continue  # skip unknown, don't fail whole batch
        duplicate = await find_recent_duplicate_payment(item.worker_id, item.amount)
        doc = {
            "id": str(uuid.uuid4()),
            "worker_id": item.worker_id,
            "worker_name": worker["name"],
            "payment_date": date,
            "amount": round(item.amount, 2),
            "payment_method": item.payment_method,
            "transaction_reference": item.transaction_reference or "",
            "notes": item.notes or "Batch payout",
            "created_by": u["id"],
            "created_at": now_utc().isoformat(),
        }
        if body.client_sync_id:
            doc["batch_sync_id"] = body.client_sync_id
        await db.payments.insert_one(doc.copy())
        await audit(u, "create", "payment", doc["id"], None, doc)
        wu = await db.users.find_one({"worker_id": item.worker_id}, {"_id": 0, "id": 1})
        if wu:
            await add_notification(wu["id"], "Payment Added", f"Payment of ₹{doc['amount']} recorded.")
        created.append(clean(doc))
        if duplicate:
            warnings.append(
                f"{worker['name']}: another ₹{doc['amount']} payment was recorded at {formatted_ist(duplicate['created_at'])}."
            )
    return {"created": len(created), "payments": created, "warnings": warnings}


# --- Payment Requests ------------------------------------------------------
# Worker → request → admin reviews/approves/rejects → admin records the actual
# payment (today: their usual method — UPI/cash/bank transfer done outside the
# app, exactly like every other payment here). No automated bank transfer
# happens anywhere in this flow — that's a deliberate choice, not a gap: real
# money movement needs a live payout-API integration (e.g. IndusInd's Transfer
# Payment API) wired in on top of this once that access is in place, with a
# human still reviewing before anything executes.
@api.post("/payment-requests")
async def create_payment_request(body: PaymentRequestIn, u=Depends(require_role(Role.worker))):
    if not u.get("worker_id"):
        raise HTTPException(400, "Only workers can request payment")
    doc = {
        "id": str(uuid.uuid4()),
        "worker_id": u["worker_id"],
        "worker_name": u.get("name"),
        "amount": round(body.amount, 2),
        "note": body.note or "",
        "status": PaymentRequestStatus.pending.value,
        "remarks": "",
        "payment_id": None,
        "created_at": now_utc().isoformat(),
    }
    await db.payment_requests.insert_one(doc.copy())
    async for admin in db.users.find({"role": Role.admin.value}, {"_id": 0, "id": 1}):
        await add_notification(admin["id"], "Payment Request", f"{u.get('name')} requested ₹{doc['amount']}.")
    return clean(doc)


@api.get("/payment-requests")
async def list_payment_requests(u=Depends(get_user), status_f: Optional[str] = Query(None, alias="status")):
    q = {}
    if u["role"] == Role.worker.value:
        q["worker_id"] = u.get("worker_id")
    if status_f:
        q["status"] = status_f
    return await db.payment_requests.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)


@api.delete("/payment-requests/{rid}")
async def cancel_payment_request(rid: str, u=Depends(require_role(Role.worker))):
    old = await db.payment_requests.find_one({"id": rid}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Not found")
    if old["worker_id"] != u.get("worker_id"):
        raise HTTPException(403, "Not your request")
    if old["status"] != PaymentRequestStatus.pending.value:
        raise HTTPException(400, "Only a Pending request can be cancelled")
    await db.payment_requests.delete_one({"id": rid})
    await audit(u, "delete", "payment_request", rid, old, None)
    return {"ok": True}


@api.post("/payment-requests/{rid}/approve")
async def approve_payment_request(rid: str, body: PaymentRequestDecision, u=Depends(require_role(Role.admin))):
    old = await db.payment_requests.find_one({"id": rid}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Not found")
    if old["status"] != PaymentRequestStatus.pending.value:
        raise HTTPException(400, "Only a Pending request can be approved")
    upd = {
        "status": PaymentRequestStatus.approved.value, "remarks": body.remarks or "",
        "decided_by": u["id"], "decided_by_name": u.get("name") or u.get("username"),
        "decided_at": now_utc().isoformat(),
    }
    await db.payment_requests.update_one({"id": rid}, {"$set": upd})
    await audit(u, "payment_request-approve", "payment_request", rid, old, upd)
    wu = await db.users.find_one({"worker_id": old["worker_id"]}, {"_id": 0, "id": 1})
    if wu:
        await add_notification(wu["id"], "Payment Request Approved", f"Your request for ₹{old['amount']} was approved.")
    return {**old, **upd}


@api.post("/payment-requests/{rid}/reject")
async def reject_payment_request(rid: str, body: PaymentRequestDecision, u=Depends(require_role(Role.admin))):
    if not (body.remarks or "").strip():
        raise HTTPException(400, "A reason is required to reject a payment request")
    old = await db.payment_requests.find_one({"id": rid}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Not found")
    if old["status"] != PaymentRequestStatus.pending.value:
        raise HTTPException(400, "Only a Pending request can be rejected")
    upd = {
        "status": PaymentRequestStatus.rejected.value, "remarks": body.remarks,
        "decided_by": u["id"], "decided_by_name": u.get("name") or u.get("username"),
        "decided_at": now_utc().isoformat(),
    }
    await db.payment_requests.update_one({"id": rid}, {"$set": upd})
    await audit(u, "payment_request-reject", "payment_request", rid, old, upd)
    wu = await db.users.find_one({"worker_id": old["worker_id"]}, {"_id": 0, "id": 1})
    if wu:
        await add_notification(wu["id"], "Payment Request Rejected", f"Your request for ₹{old['amount']} was rejected: {body.remarks}")
    return {**old, **upd}


@api.post("/payment-requests/{rid}/mark-paid")
async def mark_payment_request_paid(rid: str, body: PaymentRequestPayIn, u=Depends(require_role(Role.admin))):
    """Records the actual payment against an Approved request — the same manual
    entry as every other payment in this app, just pre-linked back to the
    request that prompted it. Idempotency-guarded like every other payment."""
    old = await db.payment_requests.find_one({"id": rid}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Not found")
    if old["status"] != PaymentRequestStatus.approved.value:
        raise HTTPException(400, "Only an Approved request can be marked paid")
    if body.client_sync_id:
        existing = await db.payments.find_one({"client_sync_id": body.client_sync_id}, {"_id": 0})
        if existing:
            return existing
    worker = await db.workers.find_one({"id": old["worker_id"]}, {"_id": 0})
    if not worker:
        raise HTTPException(404, "Worker not found")
    amount = round(body.amount if body.amount is not None else old["amount"], 2)
    duplicate = await find_recent_duplicate_payment(old["worker_id"], amount)
    payment_doc = {
        "id": str(uuid.uuid4()),
        "worker_id": old["worker_id"],
        "worker_name": worker["name"],
        "payment_date": now_utc().date().isoformat(),
        "amount": amount,
        "payment_method": body.payment_method,
        "transaction_reference": body.transaction_reference or "",
        "notes": body.notes or f"Payment request {rid}",
        "created_by": u["id"],
        "created_at": now_utc().isoformat(),
        "excel_sync_status": await _initial_excel_sync_status(),
        "excel_sync_attempts": 0,
        "excel_last_synced_at": None,
        "excel_last_error": None,
        "excel_entry_id": None,
    }
    if body.client_sync_id:
        payment_doc["client_sync_id"] = body.client_sync_id
    await db.payments.insert_one(payment_doc.copy())
    await audit(u, "create", "payment", payment_doc["id"], None, payment_doc)
    req_upd = {"status": PaymentRequestStatus.paid.value, "payment_id": payment_doc["id"]}
    await db.payment_requests.update_one({"id": rid}, {"$set": req_upd})
    await audit(u, "payment_request-paid", "payment_request", rid, old, req_upd)
    wu = await db.users.find_one({"worker_id": old["worker_id"]}, {"_id": 0, "id": 1})
    if wu:
        await add_notification(wu["id"], "Payment Sent", f"₹{amount} has been recorded as paid for your request.")
    result = {**old, **req_upd, "payment": clean(payment_doc)}
    if duplicate:
        result["duplicate_warning"] = (
            f"Another ₹{amount} payment to {worker['name']} was recorded at "
            f"{formatted_ist(duplicate['created_at'])}. Please double-check this isn't a repeat."
        )
    return result


# --- Excel Integration (Microsoft Graph / OneDrive) -----------------------
# Admin-only throughout. Workers never see any endpoint or setting in this
# section — enforced the same way as everywhere else (require_role(Role.admin)).
def _fy_label(d: Optional[datetime] = None) -> str:
    d = d or now_utc()
    y = d.year
    if d.month < 4:
        y -= 1
    return f"FY {y % 100:02d}-{(y + 1) % 100:02d}"


async def _get_excel_access_token() -> str:
    cfg = await db.excel_integration.find_one({"id": "global"}, {"_id": 0})
    if not cfg or not cfg.get("refresh_token_enc"):
        raise HTTPException(400, "Excel Integration is not connected")
    expires_at = cfg.get("token_expires_at")
    fresh_enough = False
    if expires_at:
        try:
            fresh_enough = datetime.fromisoformat(expires_at) - now_utc() > timedelta(minutes=5)
        except ValueError:
            fresh_enough = False
    if fresh_enough:
        token = decrypt_secret(cfg.get("access_token_enc"))
        if token:
            return token
    refresh_token = decrypt_secret(cfg.get("refresh_token_enc"))
    client_secret = decrypt_secret(cfg.get("client_secret_enc"))
    if not refresh_token or not client_secret:
        raise HTTPException(400, "Excel Integration needs to be reconnected")
    try:
        tokens = await excel_sync.refresh_access_token(cfg["tenant_id"], cfg["client_id"], client_secret, refresh_token)
    except excel_sync.GraphAPIError as e:
        raise HTTPException(502, f"Could not refresh Microsoft Graph token: {e}")
    new_expiry = (now_utc() + timedelta(seconds=tokens.get("expires_in", 3600))).isoformat()
    await db.excel_integration.update_one({"id": "global"}, {"$set": {
        "access_token_enc": encrypt_secret(tokens["access_token"]),
        "refresh_token_enc": encrypt_secret(tokens.get("refresh_token") or refresh_token),
        "token_expires_at": new_expiry,
    }})
    return tokens["access_token"]


async def _get_excel_store():
    cfg = await db.excel_integration.find_one({"id": "global"}, {"_id": 0})
    if not cfg or not cfg.get("enabled") or not cfg.get("drive_item_id"):
        raise HTTPException(400, "Excel Integration is not configured/connected")
    return excel_sync.GraphExcelStore(cfg["drive_item_id"], _get_excel_access_token)


@api.get("/excel-integration/settings")
async def get_excel_integration_settings(u=Depends(require_role(Role.admin))):
    cfg = await db.excel_integration.find_one({"id": "global"}, {"_id": 0}) or {}
    return {
        "enabled": cfg.get("enabled", False),
        "tenant_id": cfg.get("tenant_id"),
        "client_id": cfg.get("client_id"),
        "workbook_file_path": cfg.get("workbook_file_path"),
        "redirect_uri": cfg.get("redirect_uri"),
        "connected": bool(cfg.get("refresh_token_enc")),
        "workbook_name": cfg.get("workbook_name"),
        "last_connection_test_at": cfg.get("last_connection_test_at"),
        "last_connection_test_ok": cfg.get("last_connection_test_ok"),
        "last_connection_test_message": cfg.get("last_connection_test_message"),
        "node_letter": excel_sync.NODE_LETTER,
    }


@api.put("/excel-integration/settings")
async def update_excel_integration_settings(body: ExcelIntegrationSettingsIn, u=Depends(require_role(Role.admin))):
    old = await db.excel_integration.find_one({"id": "global"}, {"_id": 0})
    upd = {
        "tenant_id": body.tenant_id.strip() or "common",
        "client_id": body.client_id.strip(),
        "client_secret_enc": encrypt_secret(body.client_secret),
        "workbook_file_path": body.workbook_file_path.strip(),
        "redirect_uri": body.redirect_uri.strip(),
        "updated_by": u["id"],
        "updated_at": now_utc().isoformat(),
    }
    await db.excel_integration.update_one({"id": "global"}, {"$set": upd}, upsert=True)
    safe_old = {k: v for k, v in (old or {}).items() if not k.endswith("_enc")}
    safe_new = {k: v for k, v in upd.items() if not k.endswith("_enc")}
    await audit(u, "update", "excel_integration_settings", "global", safe_old, safe_new)
    return {"ok": True}


@api.get("/excel-integration/oauth/start")
async def excel_oauth_start(u=Depends(require_role(Role.admin))):
    cfg = await db.excel_integration.find_one({"id": "global"}, {"_id": 0})
    if not cfg or not cfg.get("client_id") or not cfg.get("redirect_uri") or not cfg.get("client_secret_enc"):
        raise HTTPException(400, "Set Tenant ID, Client ID, Client Secret and Redirect URI first")
    state = secrets.token_urlsafe(24)
    await db.excel_integration.update_one({"id": "global"}, {"$set": {
        "oauth_state": state,
        "oauth_state_expires": (now_utc() + timedelta(minutes=10)).isoformat(),
    }})
    params = {
        "client_id": cfg["client_id"],
        "response_type": "code",
        "redirect_uri": cfg["redirect_uri"],
        "response_mode": "query",
        "scope": "offline_access Files.ReadWrite",
        "state": state,
        # Forces Microsoft to treat this as a personal-account sign-in and skip
        # home-realm discovery — without this, an individual who also owns an
        # Azure AD tenant (e.g. from registering this very app) can get
        # intermittently routed into their own org's "work account only" login
        # experience even though the app itself allows personal accounts.
        "domain_hint": "consumers",
    }
    url = f"https://login.microsoftonline.com/{cfg['tenant_id']}/oauth2/v2.0/authorize?{urlencode(params)}"
    return {"url": url}


@api.get("/excel-integration/oauth/callback")
async def excel_oauth_callback(code: Optional[str] = None, state: Optional[str] = None,
                                error: Optional[str] = None, error_description: Optional[str] = None):
    cfg = await db.excel_integration.find_one({"id": "global"}, {"_id": 0})
    if error:
        return RedirectResponse(url=f"/?excel_connected=0&msg={quote(error_description or error)}")
    if not cfg or not state or state != cfg.get("oauth_state"):
        return RedirectResponse(url="/?excel_connected=0&msg=" + quote("Invalid or expired sign-in attempt. Please retry."))
    expires = cfg.get("oauth_state_expires")
    if expires and datetime.fromisoformat(expires) < now_utc():
        return RedirectResponse(url="/?excel_connected=0&msg=" + quote("This sign-in link expired. Please retry."))
    client_secret = decrypt_secret(cfg.get("client_secret_enc"))
    try:
        tokens = await excel_sync.exchange_code_for_tokens(
            cfg["tenant_id"], cfg["client_id"], client_secret, code, cfg["redirect_uri"]
        )
    except excel_sync.GraphAPIError as e:
        return RedirectResponse(url="/?excel_connected=0&msg=" + quote(f"Microsoft token exchange failed: {e}"))
    upd = {
        "access_token_enc": encrypt_secret(tokens["access_token"]),
        "refresh_token_enc": encrypt_secret(tokens.get("refresh_token")),
        "token_expires_at": (now_utc() + timedelta(seconds=tokens.get("expires_in", 3600))).isoformat(),
        "oauth_state": None, "oauth_state_expires": None,
    }

    async def _token_getter():
        return tokens["access_token"]

    try:
        file_info = await excel_sync.resolve_file_by_path(_token_getter, cfg["workbook_file_path"])
        upd["drive_item_id"] = file_info["id"]
        upd["workbook_name"] = file_info.get("name")
        upd["enabled"] = True
        upd["last_connection_test_at"] = now_utc().isoformat()
        upd["last_connection_test_ok"] = True
        upd["last_connection_test_message"] = "Connected"
        await db.excel_integration.update_one({"id": "global"}, {"$set": upd})
        try:
            store = excel_sync.GraphExcelStore(upd["drive_item_id"], _token_getter)
            await excel_sync.ensure_tser_node_p_rows(store, _fy_label())
        except excel_sync.GraphAPIError:
            pass  # connection itself succeeded; TSER setup can be retried by the sync loop
        return RedirectResponse(url="/?excel_connected=1")
    except excel_sync.GraphAPIError as e:
        upd["enabled"] = False
        upd["last_connection_test_ok"] = False
        upd["last_connection_test_message"] = f"Signed in, but couldn't find the workbook at that path: {e}"
        await db.excel_integration.update_one({"id": "global"}, {"$set": upd})
        return RedirectResponse(url="/?excel_connected=0&msg=" + quote(upd["last_connection_test_message"]))


@api.post("/excel-integration/test-connection")
async def test_excel_connection(u=Depends(require_role(Role.admin))):
    try:
        store = await _get_excel_store()
        await store.get_headers("TW")  # cheapest real round-trip that proves read access
        msg = "Connected — workbook is reachable and the TW table is readable."
        ok = True
    except (HTTPException, excel_sync.GraphAPIError) as e:
        ok = False
        msg = e.detail if isinstance(e, HTTPException) else str(e)
    await db.excel_integration.update_one({"id": "global"}, {"$set": {
        "last_connection_test_at": now_utc().isoformat(),
        "last_connection_test_ok": ok,
        "last_connection_test_message": msg,
    }})
    return {"ok": ok, "message": msg}


@api.post("/excel-integration/disconnect")
async def disconnect_excel(u=Depends(require_role(Role.admin))):
    old = await db.excel_integration.find_one({"id": "global"}, {"_id": 0})
    await db.excel_integration.update_one({"id": "global"}, {"$set": {
        "enabled": False, "access_token_enc": None, "refresh_token_enc": None,
        "token_expires_at": None, "drive_item_id": None, "workbook_name": None,
    }})
    await audit(u, "disconnect", "excel_integration_settings", "global", old, {"enabled": False})
    return {"ok": True}


# --- Excel Master Mapping — worker/site/door-type <-> M_PARTY/M_SITE/M_ITEM
_MAPPING_COLLECTIONS = {
    "worker": {
        "coll": "workers", "excel_id_field": "excel_party_id", "excel_table": "TP",
        "excel_id_col": "Party ID", "excel_legacy_col": "Legacy ID", "app_legacy_field": "employee_id",
        "excel_name_col": "Party Name", "app_name_field": "name",
    },
    "site": {
        "coll": "sites", "excel_id_field": "excel_site_id", "excel_table": "TS",
        "excel_id_col": "Site ID", "excel_legacy_col": "Legacy ID", "app_legacy_field": None,
        "excel_name_col": "Site Name", "app_name_field": "site_name",
    },
    "door_type": {
        "coll": "door_types", "excel_id_field": "excel_item_id", "excel_table": "TI",
        "excel_id_col": "Item ID", "excel_legacy_col": "Legacy Key", "app_legacy_field": None,
        "excel_name_col": "Work Category / Item Name", "app_name_field": "name",
    },
}


@api.get("/excel-integration/mapping/{kind}")
async def excel_mapping_list(kind: str, u=Depends(require_role(Role.admin))):
    cfg = _MAPPING_COLLECTIONS.get(kind)
    if not cfg:
        raise HTTPException(404, "Unknown mapping kind")
    store = await _get_excel_store()
    excel_rows = await store.get_rows(cfg["excel_table"])
    app_docs = await db[cfg["coll"]].find({}, {"_id": 0}).to_list(1000)

    matched, unresolved = [], []
    for doc in app_docs:
        if doc.get(cfg["excel_id_field"]) or doc.get("excel_match_status") in (
            ExcelMatchStatus.matched.value, ExcelMatchStatus.create_pending.value, ExcelMatchStatus.created.value,
        ):
            matched.append(doc)
            continue
        # Decision #1: exact Legacy ID / employee_id match auto-resolves, no confirm needed
        auto_row = None
        if cfg["app_legacy_field"] and doc.get(cfg["app_legacy_field"]):
            wanted = str(doc[cfg["app_legacy_field"]]).strip().lower()
            candidates = [
                r for r in excel_rows
                if str(r["values"].get(cfg["excel_legacy_col"]) or "").strip().lower() == wanted
            ]
            if len(candidates) == 1:
                auto_row = candidates[0]
        if auto_row:
            new_excel_id = auto_row["values"].get(cfg["excel_id_col"])
            await db[cfg["coll"]].update_one({"id": doc["id"]}, {"$set": {
                cfg["excel_id_field"]: new_excel_id, "excel_match_status": ExcelMatchStatus.matched.value,
            }})
            doc[cfg["excel_id_field"]] = new_excel_id
            doc["excel_match_status"] = ExcelMatchStatus.matched.value
            matched.append(doc)
            continue
        # Everything else — unmatched or ambiguous — goes to the one-time confirm queue.
        # Never auto-created (decision #1: "do not silently create duplicate workers").
        app_name = str(doc.get(cfg["app_name_field"]) or "").strip().lower()
        suggestions = [
            {"excel_id": r["values"].get(cfg["excel_id_col"]), "name": r["values"].get(cfg["excel_name_col"])}
            for r in excel_rows
            if app_name and app_name in str(r["values"].get(cfg["excel_name_col"]) or "").strip().lower()
        ][:5]
        unresolved.append({**doc, "suggestions": suggestions})
    return {"matched_count": len(matched), "unresolved": unresolved}


@api.post("/excel-integration/mapping/{kind}/{app_id}/confirm")
async def excel_mapping_confirm(kind: str, app_id: str, body: ExcelMappingConfirmIn, u=Depends(require_role(Role.admin))):
    cfg = _MAPPING_COLLECTIONS.get(kind)
    if not cfg:
        raise HTTPException(404, "Unknown mapping kind")
    doc = await db[cfg["coll"]].find_one({"id": app_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Not found")
    if body.create_new:
        # Decision #2: auto-create the master row — but only after this explicit,
        # one-click admin confirmation that no Excel equivalent exists.
        upd = {"excel_match_status": ExcelMatchStatus.create_pending.value}
    elif body.excel_id:
        upd = {cfg["excel_id_field"]: body.excel_id, "excel_match_status": ExcelMatchStatus.matched.value}
        if kind == "door_type":
            store = await _get_excel_store()
            row = await excel_sync.find_row_by_id(store, "TI", "Item ID", body.excel_id)
            if row:
                upd["excel_work_category"] = row["values"].get("Work Category / Item Name")
                upd["excel_specification"] = row["values"].get("Specification")
    else:
        raise HTTPException(400, "Provide either excel_id or create_new")
    await db[cfg["coll"]].update_one({"id": app_id}, {"$set": upd})
    await audit(u, "excel-mapping-confirm", kind, app_id, doc, upd)
    return {"ok": True}


async def _ensure_excel_master_row(kind: str, app_doc: dict, store) -> str:
    """Returns the Excel ID for this record, auto-creating the master row only
    when the admin has already confirmed (create_pending) it has no Excel
    equivalent. Raises if it's still Unresolved — those must be resolved once
    in Excel Master Mapping first (never silently created)."""
    cfg = _MAPPING_COLLECTIONS[kind]
    status = app_doc.get("excel_match_status")
    existing_id = app_doc.get(cfg["excel_id_field"])
    if existing_id and status in (ExcelMatchStatus.matched.value, ExcelMatchStatus.created.value):
        return existing_id
    if status != ExcelMatchStatus.create_pending.value:
        raise RuntimeError(
            f"{kind} '{app_doc.get(cfg['app_name_field'])}' isn't linked to Excel yet — "
            f"resolve it in Excel Master Mapping first."
        )
    new_id = await excel_sync.mint_master_id(store, kind)
    row = {cfg["excel_id_col"]: new_id, cfg["excel_name_col"]: app_doc.get(cfg["app_name_field"]), "Active": "Yes"}
    if kind == "worker":
        row["Party Type"] = "Worker"
        row["Phone"] = app_doc.get("mobile")
        row["Legacy ID"] = app_doc.get("employee_id")
    elif kind == "site":
        row["Site Address"] = app_doc.get("address")
    elif kind == "door_type":
        row["Item Type"] = "Service"
    await store.add_row(cfg["excel_table"], row)
    await db[cfg["coll"]].update_one({"id": app_doc["id"]}, {"$set": {
        cfg["excel_id_field"]: new_id, "excel_match_status": ExcelMatchStatus.created.value,
    }})
    return new_id


# --- Excel sync engine — idempotent (entry_id / excel_work_ids / excel_entry_id
# are the upsert keys, never re-derived from a name match) -----------------
async def _sync_work_report_now(report_id: str) -> None:
    report = await db.work_reports.find_one({"id": report_id}, {"_id": 0})
    if not report:
        return
    try:
        store = await _get_excel_store()
        worker = await db.workers.find_one({"id": report["worker_id"]}, {"_id": 0})
        site = await db.sites.find_one({"id": report["site_id"]}, {"_id": 0})
        if not worker or not site:
            raise RuntimeError("Worker or site no longer exists")
        worker_excel_id = await _ensure_excel_master_row("worker", worker, store)
        site_excel_id = await _ensure_excel_master_row("site", site, store)

        fy = _fy_label()
        work_ids = list(report.get("excel_work_ids") or [])
        rejected = report.get("approval_status") == ApprovalStatus.rejected.value
        for i, item in enumerate(report.get("door_breakdown", [])):
            door_type = await db.door_types.find_one({"id": item["door_type_id"]}, {"_id": 0})
            item_excel_id = await _ensure_excel_master_row("door_type", door_type, store) if door_type else ""
            row = {
                "Date": report["work_date"],
                "Month": datetime.fromisoformat(report["work_date"]).strftime("%b-%Y"),
                "Site ID": site_excel_id,
                "Site Name": site["site_name"],
                "Item ID": item_excel_id,
                "Work Category": (door_type or {}).get("excel_work_category") or item.get("door_type_name"),
                "Specification": (door_type or {}).get("excel_specification") or "",
                "Worker ID": worker_excel_id,
                "Worker Name": worker["name"],
                "Qty": item["count"],
                "Cost Rate": item["rate"],
                "Cost Amount": item["subtotal"],
                "Billing Rate": 0,
                "Billable Amount": 0,
                "Work Status": "Completed",
                "Billed?": "No",
                "Data Check": "APP: rejected after sync — verify before billing" if rejected else "",
                "Remarks": f"[App {report['entry_id']}] {report.get('notes') or ''}".strip(),
            }
            if i < len(work_ids):
                existing = await excel_sync.find_row_by_id(store, "TW", "Work ID", work_ids[i])
                if existing:
                    await store.update_row("TW", existing["index"], row)
                    continue
                # row vanished from Excel (e.g. a human deleted it) — recreate below
            new_id = await excel_sync.mint_transactional_id(store, "WORK", fy)
            row["Work ID"] = new_id
            await store.add_row("TW", row)
            if i < len(work_ids):
                work_ids[i] = new_id
            else:
                work_ids.append(new_id)

        await db.work_reports.update_one({"id": report_id}, {
            "$set": {
                "excel_sync_status": ExcelSyncStatus.synced.value,
                "excel_last_synced_at": now_utc().isoformat(),
                "excel_last_error": None,
                "excel_work_ids": work_ids,
            },
            "$inc": {"excel_sync_attempts": 1},
        })
    except Exception as e:
        logger.warning(f"Excel sync failed for work report {report_id}: {e}")
        await db.work_reports.update_one({"id": report_id}, {
            "$set": {"excel_sync_status": ExcelSyncStatus.failed.value, "excel_last_error": str(e)[:500]},
            "$inc": {"excel_sync_attempts": 1},
        })


async def _sync_payment_now(payment_id: str) -> None:
    payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        return
    try:
        store = await _get_excel_store()
        worker = await db.workers.find_one({"id": payment["worker_id"]}, {"_id": 0})
        if not worker:
            raise RuntimeError("Worker no longer exists")
        worker_excel_id = await _ensure_excel_master_row("worker", worker, store)
        fy = _fy_label()
        row = {
            "Voucher Type": "LABPAY",
            "Date": payment["payment_date"],
            "Month": datetime.fromisoformat(payment["payment_date"]).strftime("%b-%Y"),
            "Party ID": worker_excel_id,
            "Party Name": worker["name"],
            "Amount": payment["amount"],
            "Payment Type": "Final Payment",
            "Mode": payment.get("payment_method") or "Cash",
            "UTR / Cheque / Ref": payment.get("transaction_reference") or "",
            "Narration": f"[App payment {payment['id'][:8]}] {payment.get('notes') or ''}".strip(),
        }
        existing_entry_id = payment.get("excel_entry_id")
        if existing_entry_id:
            existing = await excel_sync.find_row_by_id(store, "TL", "Entry ID", existing_entry_id)
            if existing:
                await store.update_row("TL", existing["index"], row)
            else:
                existing_entry_id = None  # row vanished — recreate below
        if not existing_entry_id:
            row["Voucher No."] = await excel_sync.mint_transactional_id(store, "LABPAY", fy)
            existing_entry_id = await excel_sync.mint_ledger_entry_id(store)
            row["Entry ID"] = existing_entry_id
            await store.add_row("TL", row)

        await db.payments.update_one({"id": payment_id}, {
            "$set": {
                "excel_sync_status": ExcelSyncStatus.synced.value,
                "excel_last_synced_at": now_utc().isoformat(),
                "excel_last_error": None,
                "excel_entry_id": existing_entry_id,
            },
            "$inc": {"excel_sync_attempts": 1},
        })
    except Exception as e:
        logger.warning(f"Excel sync failed for payment {payment_id}: {e}")
        await db.payments.update_one({"id": payment_id}, {
            "$set": {"excel_sync_status": ExcelSyncStatus.failed.value, "excel_last_error": str(e)[:500]},
            "$inc": {"excel_sync_attempts": 1},
        })


EXCEL_SYNC_MAX_ATTEMPTS = 8
EXCEL_SYNC_INTERVAL_SECONDS = 300


async def _run_excel_sync_pass(limit: int = 50) -> dict:
    pending_q = {
        "excel_sync_status": {"$in": [ExcelSyncStatus.pending.value, ExcelSyncStatus.failed.value]},
        "excel_sync_attempts": {"$lt": EXCEL_SYNC_MAX_ATTEMPTS},
    }
    reports = await db.work_reports.find(pending_q, {"_id": 0, "id": 1}).limit(limit).to_list(limit)
    for r in reports:
        await _sync_work_report_now(r["id"])
    payments = await db.payments.find(pending_q, {"_id": 0, "id": 1}).limit(limit).to_list(limit)
    for p in payments:
        await _sync_payment_now(p["id"])
    return {"work_reports_processed": len(reports), "payments_processed": len(payments)}


async def excel_sync_loop():
    """Automatic retries — decoupled from the request path so a slow/unreachable
    Graph API never blocks an admin's approve/pay action. The DB write already
    succeeded before this loop ever runs; this only ever affects sync status."""
    while True:
        try:
            await asyncio.sleep(EXCEL_SYNC_INTERVAL_SECONDS)
            cfg = await db.excel_integration.find_one({"id": "global"}, {"_id": 0, "enabled": 1})
            if not cfg or not cfg.get("enabled"):
                continue
            await _run_excel_sync_pass()
        except asyncio.CancelledError:
            break
        except Exception as e:
            logger.warning(f"Excel sync loop error (non-fatal, will retry): {e}")


@api.post("/excel-integration/sync-now")
async def excel_sync_now(u=Depends(require_role(Role.admin))):
    result = await _run_excel_sync_pass(limit=200)
    await audit(u, "sync-now", "excel_integration", "global", None, result)
    return result


@api.post("/excel-integration/retry/{entity_type}/{entity_id}")
async def excel_retry_one(entity_type: str, entity_id: str, u=Depends(require_role(Role.admin))):
    if entity_type == "work_report":
        await db.work_reports.update_one({"id": entity_id}, {"$set": {"excel_sync_status": ExcelSyncStatus.pending.value}})
        await _sync_work_report_now(entity_id)
    elif entity_type == "payment":
        await db.payments.update_one({"id": entity_id}, {"$set": {"excel_sync_status": ExcelSyncStatus.pending.value}})
        await _sync_payment_now(entity_id)
    else:
        raise HTTPException(404, "Unknown entity type")
    return {"ok": True}


@api.get("/excel-integration/sync-status")
async def excel_sync_status_endpoint(u=Depends(require_role(Role.admin))):
    async def counts(coll):
        rows = await db[coll].aggregate([{"$group": {"_id": "$excel_sync_status", "n": {"$sum": 1}}}]).to_list(10)
        return {(r["_id"] or "Unknown"): r["n"] for r in rows}

    work_counts = await counts("work_reports")
    payment_counts = await counts("payments")
    last_success = await db.work_reports.find(
        {"excel_sync_status": ExcelSyncStatus.synced.value}, {"_id": 0, "excel_last_synced_at": 1},
    ).sort("excel_last_synced_at", -1).limit(1).to_list(1)
    failed_reports = await db.work_reports.find(
        {"excel_sync_status": ExcelSyncStatus.failed.value},
        {"_id": 0, "id": 1, "entry_id": 1, "worker_name": 1, "site_name": 1, "excel_last_error": 1, "excel_sync_attempts": 1},
    ).to_list(200)
    failed_payments = await db.payments.find(
        {"excel_sync_status": ExcelSyncStatus.failed.value},
        {"_id": 0, "id": 1, "worker_name": 1, "amount": 1, "excel_last_error": 1, "excel_sync_attempts": 1},
    ).to_list(200)
    workers = await db.workers.find(
        {}, {"_id": 0, "id": 1, "name": 1, "first_valid_entry_at": 1, "automation_started_at": 1, "excel_match_status": 1},
    ).to_list(1000)
    return {
        "work_reports": work_counts,
        "payments": payment_counts,
        "last_successful_sync_at": (last_success[0]["excel_last_synced_at"] if last_success else None),
        "failed_work_reports": failed_reports,
        "failed_payments": failed_payments,
        "workers": [
            {
                "worker_id": w["id"], "name": w["name"],
                "first_valid_entry_at": w.get("first_valid_entry_at"),
                "automation_started_at": w.get("automation_started_at"),
                "excel_match_status": w.get("excel_match_status") or ExcelMatchStatus.unresolved.value,
            }
            for w in workers
        ],
    }


# --- Payment / Banking API (kept fully separate from Excel Integration) ---
@api.get("/banking-integration/settings")
async def get_banking_settings(u=Depends(require_role(Role.admin))):
    cfg = await db.banking_integration.find_one({"id": "global"}, {"_id": 0}) or {}
    return {
        "enabled": cfg.get("enabled", False),
        "provider_name": cfg.get("provider_name", ""),
        "api_base_url": cfg.get("api_base_url", ""),
        "environment": cfg.get("environment", "Sandbox"),
        "has_api_key": bool(cfg.get("api_key_enc")),
        "has_client_secret": bool(cfg.get("client_secret_enc")),
        "merchant_id": cfg.get("merchant_id", ""),
        "webhook_url": cfg.get("webhook_url", ""),
        "has_webhook_secret": bool(cfg.get("webhook_secret_enc")),
        "last_test_at": cfg.get("last_test_at"),
        "last_test_ok": cfg.get("last_test_ok"),
        "last_test_message": cfg.get("last_test_message"),
    }


@api.put("/banking-integration/settings")
async def update_banking_settings(body: BankingSettingsIn, u=Depends(require_role(Role.admin))):
    old = await db.banking_integration.find_one({"id": "global"}, {"_id": 0})
    upd = {
        "provider_name": body.provider_name.strip(),
        "api_base_url": body.api_base_url.strip(),
        "merchant_id": body.merchant_id or "",
        "webhook_url": body.webhook_url or "",
        "environment": body.environment if body.environment in ("Sandbox", "Production") else "Sandbox",
        "enabled": bool(body.enabled and body.api_base_url and (body.api_key or body.client_secret)),
        "updated_by": u["id"],
        "updated_at": now_utc().isoformat(),
    }
    if body.api_key:
        upd["api_key_enc"] = encrypt_secret(body.api_key)
    if body.client_id:
        upd["client_id_enc"] = encrypt_secret(body.client_id)
    if body.client_secret:
        upd["client_secret_enc"] = encrypt_secret(body.client_secret)
    if body.webhook_secret:
        upd["webhook_secret_enc"] = encrypt_secret(body.webhook_secret)
    await db.banking_integration.update_one({"id": "global"}, {"$set": upd}, upsert=True)
    safe_old = {k: v for k, v in (old or {}).items() if not k.endswith("_enc")}
    safe_new = {k: v for k, v in upd.items() if not k.endswith("_enc")}
    await audit(u, "update", "banking_integration_settings", "global", safe_old, safe_new)
    return {"ok": True}


@api.post("/banking-integration/test-connection")
async def test_banking_connection(u=Depends(require_role(Role.admin))):
    cfg = await db.banking_integration.find_one({"id": "global"}, {"_id": 0})
    if not cfg or not cfg.get("api_base_url"):
        raise HTTPException(400, "Set an API Base URL first")
    api_key = decrypt_secret(cfg.get("api_key_enc"))
    headers = {"Authorization": f"Bearer {api_key}"} if api_key else {}
    ok, msg = False, ""
    try:
        async with httpx.AsyncClient(timeout=10.0) as c:
            r = await c.get(cfg["api_base_url"], headers=headers)
        ok = r.status_code < 500
        msg = (
            f"Reached {cfg['api_base_url']} (HTTP {r.status_code}). This confirms network "
            f"reachability and that credentials were sent — it does NOT confirm the provider "
            f"accepted them, since every bank's auth handshake is different. Confirm a real "
            f"sandbox transaction with your provider's own docs before relying on this."
        )
    except httpx.HTTPError as e:
        msg = f"Could not reach {cfg['api_base_url']}: {e}"
    await db.banking_integration.update_one({"id": "global"}, {"$set": {
        "last_test_at": now_utc().isoformat(), "last_test_ok": ok, "last_test_message": msg,
    }})
    return {"ok": ok, "message": msg}


# --- Audit log -----------------------------------------------------------
@api.get("/audit-logs")
async def list_audit(
    u=Depends(require_role(Role.admin)),
    entity_type: Optional[str] = None,
    action: Optional[str] = None,
    limit: int = 200,
):
    q = {}
    if entity_type: q["entity_type"] = entity_type
    if action: q["action"] = action
    items = await db.audit_logs.find(q, {"_id": 0}).sort("created_at", -1).limit(max(1, min(limit, 500))).to_list(500)
    return items


@api.get("/search")
async def global_search(q: str = Query(..., min_length=1), u=Depends(require_role(Role.admin))):
    """One box across worker name/employee ID/mobile and site name/client/address —
    not a replacement for the per-screen searches, an addition for "type a name,
    find it wherever it lives" without knowing which screen it's on first."""
    q_re = {"$regex": re.escape(q.strip()), "$options": "i"}
    workers = await db.workers.find(
        {"$or": [{"name": q_re}, {"employee_id": q_re}, {"mobile": q_re}]}, {"_id": 0}
    ).limit(20).to_list(20)
    sites = await db.sites.find(
        {"$or": [{"site_name": q_re}, {"client_name": q_re}, {"address": q_re}]}, {"_id": 0}
    ).limit(20).to_list(20)
    return {
        "workers": [
            {"id": w["id"], "name": w["name"], "employee_id": w.get("employee_id"), "mobile": w.get("mobile"), "status": w.get("status")}
            for w in workers
        ],
        "sites": [
            {"id": s["id"], "site_name": s["site_name"], "client_name": s.get("client_name"), "address": s.get("address"), "status": s.get("status")}
            for s in sites
        ],
    }


# --- Branches ------------------------------------------------------------
class BranchIn(BaseModel):
    name: str
    code: str
    address: Optional[str] = ""
    phone: Optional[str] = ""


@api.get("/branches")
async def list_branches(u=Depends(get_user)):
    return await db.branches.find({}, {"_id": 0}).sort("name", 1).to_list(200)


@api.post("/branches")
async def create_branch(body: BranchIn, u=Depends(require_role(Role.admin))):
    doc = {"id": str(uuid.uuid4()), **body.dict(), "created_at": now_utc().isoformat()}
    await db.branches.insert_one(doc.copy())
    await audit(u, "create", "branch", doc["id"], None, doc)
    return clean(doc)


@api.put("/branches/{bid}")
async def update_branch(bid: str, body: BranchIn, u=Depends(require_role(Role.admin))):
    old = await db.branches.find_one({"id": bid}, {"_id": 0})
    if not old: raise HTTPException(404, "Branch not found")
    upd = body.dict()
    await db.branches.update_one({"id": bid}, {"$set": upd})
    await audit(u, "update", "branch", bid, old, upd)
    return {**old, **upd}


# --- Leaves --------------------------------------------------------------
class LeaveIn(BaseModel):
    start_date: str
    end_date: str
    reason: str
    leave_type: Optional[str] = "Personal"


class LeaveDecision(BaseModel):
    remarks: Optional[str] = ""


@api.post("/leaves")
async def create_leave(body: LeaveIn, u=Depends(require_role(Role.worker, Role.admin))):
    if not u.get("worker_id"):
        raise HTTPException(400, "Only workers can request leave")
    if body.end_date < body.start_date:
        raise HTTPException(400, "end_date must be >= start_date")
    # Overlap check against this worker's own still-live requests — a date range
    # overlaps [start, end] when it starts on/before our end and ends on/after our
    # start. ISO YYYY-MM-DD strings sort correctly with plain comparison.
    overlap = await db.leaves.find_one({
        "worker_id": u["worker_id"],
        "status": {"$in": [LeaveStatus.pending.value, LeaveStatus.approved.value]},
        "start_date": {"$lte": body.end_date},
        "end_date": {"$gte": body.start_date},
    }, {"_id": 0})
    if overlap:
        raise HTTPException(
            400,
            f"This overlaps your {overlap['status'].lower()} leave request for "
            f"{overlap['start_date']} → {overlap['end_date']}.",
        )
    doc = {
        "id": str(uuid.uuid4()),
        "worker_id": u["worker_id"],
        "worker_name": u.get("name"),
        "start_date": body.start_date,
        "end_date": body.end_date,
        "reason": body.reason,
        "leave_type": body.leave_type or "Personal",
        "status": LeaveStatus.pending.value,
        "remarks": "",
        "created_at": now_utc().isoformat(),
    }
    await db.leaves.insert_one(doc.copy())
    async for admin in db.users.find({"role": Role.admin.value}, {"_id": 0, "id": 1}):
        await add_notification(admin["id"], "Leave Request", f"{u.get('name')} requested leave {body.start_date} → {body.end_date}.")
    return clean(doc)


@api.get("/leaves")
async def list_leaves(u=Depends(get_user), status_f: Optional[str] = Query(None, alias="status"),
                      worker_id: Optional[str] = None):
    q = {}
    if u["role"] == Role.worker.value:
        q["worker_id"] = u.get("worker_id")
    elif worker_id:
        q["worker_id"] = worker_id
    if status_f: q["status"] = status_f
    return await db.leaves.find(q, {"_id": 0}).sort("start_date", -1).to_list(500)


async def _leave_decide(lid: str, new_status: str, remarks: str, u: dict):
    old = await db.leaves.find_one({"id": lid}, {"_id": 0})
    if not old: raise HTTPException(404, "Leave not found")
    upd = {"status": new_status, "remarks": remarks, "decided_by": u["id"],
           "decided_by_name": u.get("name") or u.get("username"),
           "decided_at": now_utc().isoformat()}
    await db.leaves.update_one({"id": lid}, {"$set": upd})
    await audit(u, f"leave-{new_status.lower()}", "leave", lid, old, upd)
    wu = await db.users.find_one({"worker_id": old["worker_id"]}, {"_id": 0, "id": 1})
    if wu:
        await add_notification(wu["id"], f"Leave {new_status}", f"Your leave {old['start_date']} → {old['end_date']} has been {new_status.lower()}.")
    return {**old, **upd}


@api.delete("/leaves/{lid}")
async def cancel_leave(lid: str, u=Depends(require_role(Role.worker))):
    """Worker withdraws their own still-Pending request — mirrors work-report withdraw."""
    old = await db.leaves.find_one({"id": lid}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Not found")
    if old["worker_id"] != u.get("worker_id"):
        raise HTTPException(403, "Not your leave request")
    if old["status"] != LeaveStatus.pending.value:
        raise HTTPException(400, "Only a Pending request can be cancelled")
    await db.leaves.delete_one({"id": lid})
    await audit(u, "delete", "leave", lid, old, None)
    return {"ok": True}


@api.post("/leaves/{lid}/approve")
async def approve_leave(lid: str, body: LeaveDecision, u=Depends(require_role(Role.admin))):
    return await _leave_decide(lid, LeaveStatus.approved.value, body.remarks or "", u)


@api.post("/leaves/{lid}/reject")
async def reject_leave(lid: str, body: LeaveDecision, u=Depends(require_role(Role.admin))):
    return await _leave_decide(lid, LeaveStatus.rejected.value, body.remarks or "", u)


# --- GST Invoices --------------------------------------------------------
class InvoiceItemIn(BaseModel):
    description: str
    quantity: float = Field(gt=0)
    unit_price: float = Field(ge=0)
    hsn_code: Optional[str] = ""


class InvoiceIn(BaseModel):
    site_id: Optional[str] = None
    client_name: str
    client_gstin: Optional[str] = ""
    client_address: Optional[str] = ""
    invoice_date: Optional[str] = None
    due_date: Optional[str] = None
    items: List[InvoiceItemIn]
    gst_rate: float = Field(default=18.0, ge=0, le=100)  # percent
    notes: Optional[str] = ""


async def _next_invoice_number() -> str:
    year = now_utc().strftime("%y")
    counter_key = f"inv-{year}"
    r = await db.counters.find_one_and_update(
        {"id": counter_key},
        {"$inc": {"value": 1}},
        upsert=True,
        return_document=True,
    )
    n = (r or {}).get("value", 1)
    return f"JE/{year}/{n:04d}"


@api.post("/invoices")
async def create_invoice(body: InvoiceIn, u=Depends(require_role(Role.admin))):
    if not body.items:
        raise HTTPException(400, "At least one line item required")
    subtotal = round(sum(it.quantity * it.unit_price for it in body.items), 2)
    gst_amount = round(subtotal * (body.gst_rate / 100.0), 2)
    total = round(subtotal + gst_amount, 2)
    number = await _next_invoice_number()
    doc = {
        "id": str(uuid.uuid4()),
        "invoice_number": number,
        "site_id": body.site_id,
        "client_name": body.client_name,
        "client_gstin": body.client_gstin or "",
        "client_address": body.client_address or "",
        "invoice_date": body.invoice_date or now_utc().date().isoformat(),
        "due_date": body.due_date or "",
        "items": [it.dict() for it in body.items],
        "subtotal": subtotal,
        "gst_rate": body.gst_rate,
        "gst_amount": gst_amount,
        "total": total,
        "notes": body.notes or "",
        "status": "Unpaid",
        "created_by": u["id"],
        "created_at": now_utc().isoformat(),
    }
    await db.invoices.insert_one(doc.copy())
    await audit(u, "create", "invoice", doc["id"], None, doc)
    return clean(doc)


@api.put("/invoices/{iid}")
async def update_invoice(iid: str, body: InvoiceIn, u=Depends(require_role(Role.admin))):
    old = await db.invoices.find_one({"id": iid}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Not found")
    if not body.items:
        raise HTTPException(400, "At least one line item required")
    # totals are always server-recomputed from the edited line items, never trusted from the client
    subtotal = round(sum(it.quantity * it.unit_price for it in body.items), 2)
    gst_amount = round(subtotal * (body.gst_rate / 100.0), 2)
    total = round(subtotal + gst_amount, 2)
    upd = {
        "site_id": body.site_id,
        "client_name": body.client_name,
        "client_gstin": body.client_gstin or "",
        "client_address": body.client_address or "",
        "invoice_date": body.invoice_date or old["invoice_date"],
        "due_date": body.due_date or "",
        "items": [it.dict() for it in body.items],
        "subtotal": subtotal,
        "gst_rate": body.gst_rate,
        "gst_amount": gst_amount,
        "total": total,
        "notes": body.notes or "",
    }
    await db.invoices.update_one({"id": iid}, {"$set": upd})
    await audit(u, "update", "invoice", iid, old, upd)
    return {**old, **upd}


@api.post("/invoices/{iid}/mark-paid")
async def mark_invoice_paid(iid: str, u=Depends(require_role(Role.admin))):
    old = await db.invoices.find_one({"id": iid}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Not found")
    await db.invoices.update_one({"id": iid}, {"$set": {"status": "Paid"}})
    await audit(u, "update", "invoice", iid, old, {"status": "Paid"})
    return {**old, "status": "Paid"}


@api.post("/invoices/{iid}/mark-unpaid")
async def mark_invoice_unpaid(iid: str, u=Depends(require_role(Role.admin))):
    old = await db.invoices.find_one({"id": iid}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Not found")
    await db.invoices.update_one({"id": iid}, {"$set": {"status": "Unpaid"}})
    await audit(u, "update", "invoice", iid, old, {"status": "Unpaid"})
    return {**old, "status": "Unpaid"}


@api.delete("/invoices/{iid}")
async def delete_invoice(iid: str, u=Depends(require_role(Role.admin))):
    old = await db.invoices.find_one({"id": iid}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Not found")
    await db.invoices.delete_one({"id": iid})
    await audit(u, "delete", "invoice", iid, old, None)
    return {"ok": True}


@api.get("/invoices")
async def list_invoices(u=Depends(require_role(Role.admin, Role.client))):
    q = {}
    if u["role"] == Role.client.value:
        q["site_id"] = {"$in": u.get("site_ids", []) or []}
    return await db.invoices.find(q, {"_id": 0}).sort("invoice_date", -1).to_list(500)


@api.get("/invoices/{iid}")
async def get_invoice(iid: str, u=Depends(require_role(Role.admin, Role.client))):
    inv = await db.invoices.find_one({"id": iid}, {"_id": 0})
    if not inv: raise HTTPException(404, "Not found")
    if u["role"] == Role.client.value and inv.get("site_id") not in (u.get("site_ids") or []):
        raise HTTPException(403, "Forbidden")
    return inv


def _pdf_text(value) -> str:
    """Escape free text before it goes into a ReportLab Paragraph() — Paragraph
    interprets a subset of XML/HTML-like markup in its input, so an unescaped
    '<' or '&' in admin-entered text (a client name, notes, etc.) throws a parse
    error and crashes that document's generation instead of just rendering oddly."""
    return _xml_escape(str(value)) if value else ""


def _build_invoice_pdf(inv: dict, company: dict) -> bytes:
    from reportlab.lib import colors as rlc
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=32, rightMargin=32, topMargin=32, bottomMargin=32)
    styles = getSampleStyleSheet()
    brand = ParagraphStyle("brand", parent=styles["Heading1"], textColor=rlc.HexColor("#423226"))
    story = [
        Paragraph(_pdf_text(company.get("company_name", "JONAH ENTERPRISES")), brand),
        Paragraph(_pdf_text(company.get("address")), styles["Normal"]),
        Paragraph(f"{_pdf_text(company.get('mobile'))}   {_pdf_text(company.get('email'))}", styles["Normal"]),
        Spacer(1, 10),
        Paragraph(f"<b>TAX INVOICE — {_pdf_text(inv['invoice_number'])}</b>", styles["Heading2"]),
        Paragraph(f"Date: {_pdf_text(inv['invoice_date'])}   Due: {_pdf_text(inv.get('due_date')) or '-'}", styles["Normal"]),
        Spacer(1, 8),
        Paragraph(f"<b>Bill to:</b> {_pdf_text(inv['client_name'])}", styles["Normal"]),
        Paragraph(_pdf_text(inv.get("client_address")), styles["Normal"]),
        Paragraph(f"GSTIN: {_pdf_text(inv.get('client_gstin')) or '-'}", styles["Normal"]),
        Spacer(1, 10),
    ]
    header = ["Description", "HSN", "Qty", "Unit ₹", "Amount ₹"]
    rows = [[it["description"], it.get("hsn_code") or "-", f"{it['quantity']:g}", f"{it['unit_price']:.2f}", f"{it['quantity'] * it['unit_price']:.2f}"] for it in inv["items"]]
    footer_rows = [
        ["", "", "", "Subtotal", f"{inv['subtotal']:.2f}"],
        ["", "", "", f"GST @ {inv['gst_rate']:.1f}%", f"{inv['gst_amount']:.2f}"],
        ["", "", "", "TOTAL", f"₹ {inv['total']:.2f}"],
    ]
    table_data = [header] + rows + footer_rows
    t = Table(table_data, colWidths=[220, 60, 45, 70, 90])
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rlc.HexColor("#423226")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rlc.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, len(rows)), [rlc.HexColor("#FAF9F6"), rlc.white]),
        ("GRID", (0, 0), (-1, -1), 0.25, rlc.HexColor("#CFC9BC")),
        ("BACKGROUND", (0, -1), (-1, -1), rlc.HexColor("#EADDBC")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ])
    t.setStyle(style)
    story.append(t)
    if inv.get("notes"):
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"<b>Notes:</b> {_pdf_text(inv['notes'])}", styles["Normal"]))
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


@api.get("/invoices/{iid}/pdf")
async def invoice_pdf(iid: str, token: Optional[str] = None):
    if not token: raise HTTPException(401, "Missing download token")
    await _verify_admin_download_token(token)
    inv = await db.invoices.find_one({"id": iid}, {"_id": 0})
    if not inv: raise HTTPException(404, "Not found")
    company = await _company_settings()
    data = _build_invoice_pdf(inv, company)
    fname = f"{inv['invoice_number'].replace('/', '_')}.pdf"
    return StreamingResponse(io.BytesIO(data), media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# --- Salary Slip --------------------------------------------------------
@api.get("/salary-slip/{worker_id}")
async def salary_slip(worker_id: str, month: Optional[str] = None, token: Optional[str] = None):
    """month=YYYY-MM. Auth via ?token= (download token)."""
    if not token:
        raise HTTPException(401, "Missing download token")
    payload = await _verify_download_token(token)
    # Worker-scoped tokens (minted by /worker/salary-slip-token) may only fetch their own slip.
    if payload.get("role") == Role.worker.value and payload.get("worker_id") != worker_id:
        raise HTTPException(403, "Forbidden")
    month = month or now_utc().strftime("%Y-%m")
    if not re.match(r"^\d{4}-(0[1-9]|1[0-2])$", month):
        raise HTTPException(400, "month must be in YYYY-MM format")
    worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    if not worker: raise HTTPException(404, "Worker not found")

    start = f"{month}-01"
    # naive end: 32 days later then next month's -01
    year, m = map(int, month.split("-"))
    ny, nm = (year + 1, 1) if m == 12 else (year, m + 1)
    end = f"{ny:04d}-{nm:02d}-01"

    doors = 0; total = 0.0; days = set(); sessions = 0
    async for r in db.work_reports.find(
        {"worker_id": worker_id, "approval_status": "Approved",
         "work_date": {"$gte": start, "$lt": end}}, {"_id": 0}
    ):
        doors += r.get("door_count", 0)
        total += float(r.get("total_amount", 0))
        days.add(r.get("work_date"))
    async for a in db.attendance_sessions.find(
        {"worker_id": worker_id, "check_in_time": {"$gte": start, "$lt": end}}, {"_id": 0, "id": 1}
    ):
        sessions += 1

    paid = 0.0
    async for p in db.payments.find(
        {"worker_id": worker_id, "payment_date": {"$gte": start, "$lt": end}}, {"_id": 0}
    ):
        paid += float(p.get("amount", 0))

    from reportlab.lib import colors as rlc
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    company = await _company_settings()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=32, rightMargin=32, topMargin=32, bottomMargin=32)
    styles = getSampleStyleSheet()
    brand = ParagraphStyle("brand", parent=styles["Heading1"], textColor=rlc.HexColor("#423226"))
    story = [
        Paragraph(_pdf_text(company.get("company_name", "JONAH ENTERPRISES")), brand),
        Paragraph(_pdf_text(company.get("address")), styles["Normal"]),
        Spacer(1, 10),
        Paragraph(f"<b>PAYMENT SLIP — {month}</b>", styles["Heading2"]),
        Paragraph(f"Worker: {_pdf_text(worker['name'])} ({_pdf_text(worker.get('employee_id')) or '-'})", styles["Normal"]),
        Paragraph(f"Mobile: {_pdf_text(worker.get('mobile')) or '-'}", styles["Normal"]),
        Spacer(1, 10),
    ]
    # Actual blended rate for the month, not the worker's flat default_rate — door
    # types and per-site overrides mean those can genuinely differ, and showing the
    # wrong one here would visibly disagree with the gross figure right below it.
    effective_rate = (total / doors) if doors else worker.get("default_rate", 0)
    tbl = Table([
        ["Metric", "Value"],
        ["Working Days", len(days)],
        ["Site Sessions", sessions],
        ["Doors Installed", doors],
        ["Rate Per Door", f"₹ {effective_rate:.2f}"],
        ["Gross Approved Earnings", f"₹ {total:.2f}"],
        ["Amount Paid This Month", f"₹ {paid:.2f}"],
        ["Net Balance Carried", f"₹ {total - paid:.2f}"],
    ], colWidths=[240, 220])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rlc.HexColor("#423226")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rlc.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, rlc.HexColor("#CFC9BC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rlc.HexColor("#FAF9F6"), rlc.white]),
        ("BACKGROUND", (0, -1), (-1, -1), rlc.HexColor("#EADDBC")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 20))
    story.append(Paragraph("This is a system-generated payment slip. Signature is not required.", styles["Italic"]))
    doc.build(story)
    buf.seek(0)
    fname = f"slip_{worker.get('employee_id', worker_id)}_{month}.pdf"
    return StreamingResponse(io.BytesIO(buf.getvalue()), media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# --- Client Portal -------------------------------------------------------
class ClientUserIn(BaseModel):
    username: str
    password: str = Field(min_length=8)
    display_name: str
    site_ids: List[str]


@api.post("/client-users")
async def create_client_user(body: ClientUserIn, u=Depends(require_role(Role.admin))):
    if await db.users.find_one({"username": body.username.lower().strip()}):
        raise HTTPException(400, "Username already exists")
    doc = {
        "id": str(uuid.uuid4()),
        "username": body.username.lower().strip(),
        "password_hash": pwd_ctx.hash(body.password),
        "role": Role.client.value,
        "name": body.display_name,
        "site_ids": body.site_ids,
        "disabled": False,
        "created_at": now_utc().isoformat(),
    }
    await db.users.insert_one(doc.copy())
    await audit(u, "create", "client_user", doc["id"], None, {"username": body.username, "sites": body.site_ids})
    return {"id": doc["id"], "username": doc["username"], "name": doc["name"]}


@api.get("/client/dashboard")
async def client_dashboard(u=Depends(require_role(Role.client, Role.admin))):
    site_ids = u.get("site_ids") or []
    if u["role"] == Role.admin.value:
        # admin viewing "as client" simply sees all sites
        sites = await db.sites.find({}, {"_id": 0}).to_list(200)
    else:
        sites = await db.sites.find({"id": {"$in": site_ids}}, {"_id": 0}).to_list(200)
    doors_agg = await db.work_reports.aggregate([
        {"$match": {"site_id": {"$in": [s["id"] for s in sites]}, "approval_status": "Approved"}},
        {"$group": {"_id": "$site_id", "doors": {"$sum": "$door_count"},
                    "amount": {"$sum": "$total_amount"}, "visits": {"$sum": 1}}}
    ]).to_list(1000)
    by = {a["_id"]: a for a in doors_agg}
    rows = []
    for s in sites:
        m = by.get(s["id"], {})
        rows.append({
            "site_id": s["id"], "site_name": s["site_name"], "address": s.get("address"),
            "doors": m.get("doors", 0), "amount": round(m.get("amount", 0), 2),
            "visits": m.get("visits", 0), "status": s.get("status"),
        })
    return {
        "sites": rows,
        "total_doors": sum(r["doors"] for r in rows),
        "total_amount": round(sum(r["amount"] for r in rows), 2),
        "site_count": len(rows),
    }


@api.get("/client/reports")
async def client_reports(u=Depends(require_role(Role.client, Role.admin)),
                         site_id: Optional[str] = None):
    q = {"approval_status": "Approved"}
    if u["role"] == Role.client.value:
        allowed = u.get("site_ids") or []
        if site_id:
            if site_id not in allowed: raise HTTPException(403, "Forbidden")
            q["site_id"] = site_id
        else:
            q["site_id"] = {"$in": allowed}
    elif site_id:
        q["site_id"] = site_id
    items = await db.work_reports.find(q, {"_id": 0, "before_photos": 0, "after_photos": 0}).sort("work_date", -1).limit(200).to_list(200)
    return items


# --- Charts (admin) ------------------------------------------------------
@api.get("/admin/charts")
async def admin_charts(u=Depends(require_role(Role.admin))):
    """Aggregates for dashboard charts: last-6-months doors installed,
    last-6-months payments, top workers by doors, site-wise installations."""
    now = now_utc().date()

    # Build the last 6 months (YYYY-MM keys, oldest first)
    months = []
    y, m = now.year, now.month
    for _ in range(6):
        months.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            m = 12; y -= 1
    months.reverse()
    month_start = months[0] + "-01"

    # Monthly doors (approved reports only for accounting truth) — grouped in
    # Mongo rather than pulled document-by-document into Python, consistent with
    # the top-workers/site-installations aggregations just below.
    doors_by_month = {mm: 0 for mm in months}
    doors_agg = await db.work_reports.aggregate([
        {"$match": {"approval_status": "Approved", "work_date": {"$gte": month_start}}},
        {"$group": {"_id": {"$substrCP": ["$work_date", 0, 7]}, "doors": {"$sum": "$door_count"}}},
    ]).to_list(100)
    for d in doors_agg:
        if d["_id"] in doors_by_month:
            doors_by_month[d["_id"]] = d["doors"]

    # Monthly payments
    pay_by_month = {mm: 0.0 for mm in months}
    pay_agg = await db.payments.aggregate([
        {"$match": {"payment_date": {"$gte": month_start}}},
        {"$group": {"_id": {"$substrCP": ["$payment_date", 0, 7]}, "amount": {"$sum": "$amount"}}},
    ]).to_list(100)
    for p in pay_agg:
        if p["_id"] in pay_by_month:
            pay_by_month[p["_id"]] = float(p["amount"])

    # Top workers by approved doors
    top_workers_agg = await db.work_reports.aggregate([
        {"$match": {"approval_status": "Approved"}},
        {"$group": {"_id": "$worker_id", "worker_name": {"$first": "$worker_name"},
                    "doors": {"$sum": "$door_count"}, "amount": {"$sum": "$total_amount"}}},
        {"$sort": {"doors": -1}},
        {"$limit": 6},
    ]).to_list(6)

    # Site-wise installations
    sites_agg = await db.work_reports.aggregate([
        {"$match": {"approval_status": "Approved"}},
        {"$group": {"_id": "$site_id", "site_name": {"$first": "$site_name"},
                    "doors": {"$sum": "$door_count"}, "amount": {"$sum": "$total_amount"}}},
        {"$sort": {"doors": -1}},
        {"$limit": 8},
    ]).to_list(8)

    return {
        "monthly_doors": [{"month": mm, "value": doors_by_month[mm]} for mm in months],
        "monthly_payments": [{"month": mm, "value": round(pay_by_month[mm], 2)} for mm in months],
        "top_workers": [{"worker_id": t["_id"], "worker_name": t["worker_name"], "doors": t["doors"], "amount": round(t["amount"], 2)} for t in top_workers_agg],
        "site_installations": [{"site_id": s["_id"], "site_name": s["site_name"], "doors": s["doors"], "amount": round(s["amount"], 2)} for s in sites_agg],
    }


# --- Reports & Export ----------------------------------------------------
def _rupees(n) -> str:
    try:
        return f"₹{float(n or 0):,.2f}"
    except Exception:
        return "₹0.00"


async def _rpt_worker_earnings(date_from: Optional[str], date_to: Optional[str], worker_id: Optional[str]):
    q = {"approval_status": "Approved"}
    if date_from: q.setdefault("work_date", {})["$gte"] = date_from
    if date_to: q.setdefault("work_date", {})["$lte"] = date_to
    if worker_id: q["worker_id"] = worker_id
    rows = []
    totals = {"doors": 0, "amount": 0.0}
    by_worker: dict = {}
    async for r in db.work_reports.find(q, {"_id": 0}):
        wid = r["worker_id"]
        if wid not in by_worker:
            by_worker[wid] = {"worker": r.get("worker_name", "-"), "doors": 0, "amount": 0.0}
        by_worker[wid]["doors"] += r.get("door_count", 0)
        by_worker[wid]["amount"] += float(r.get("total_amount", 0))
        totals["doors"] += r.get("door_count", 0)
        totals["amount"] += float(r.get("total_amount", 0))
    # One grouped query for every worker's all-time paid total instead of one
    # query per worker inside the loop below.
    worker_ids = list(by_worker.keys())
    paid_agg = await db.payments.aggregate([
        {"$match": {"worker_id": {"$in": worker_ids}}},
        {"$group": {"_id": "$worker_id", "total": {"$sum": "$amount"}}},
    ]).to_list(1000) if worker_ids else []
    paid_by_worker = {r["_id"]: float(r["total"]) for r in paid_agg}
    for wid, v in by_worker.items():
        paid = paid_by_worker.get(wid, 0.0)
        rows.append([v["worker"], v["doors"], _rupees(v["amount"]), _rupees(paid), _rupees(v["amount"] - paid)])
    header = ["Worker", "Doors", "Approved Amount", "Paid", "Pending"]
    footer = ["TOTAL", totals["doors"], _rupees(totals["amount"]), "", ""]
    return header, rows, footer, "Worker Earnings Report"


async def _rpt_payments(date_from, date_to, worker_id):
    q = {}
    if date_from: q.setdefault("payment_date", {})["$gte"] = date_from
    if date_to: q.setdefault("payment_date", {})["$lte"] = date_to
    if worker_id: q["worker_id"] = worker_id
    rows = []
    total = 0.0
    async for p in db.payments.find(q, {"_id": 0}).sort("payment_date", -1):
        rows.append([p.get("payment_date"), p.get("worker_name"), p.get("payment_method"),
                     p.get("transaction_reference") or "-", _rupees(p.get("amount"))])
        total += float(p.get("amount", 0))
    header = ["Date", "Worker", "Method", "Reference", "Amount"]
    footer = ["TOTAL", "", "", "", _rupees(total)]
    return header, rows, footer, "Payments Report"


async def _rpt_attendance(date_from, date_to, worker_id, site_id):
    q = {}
    if date_from: q.setdefault("check_in_time", {})["$gte"] = date_from
    if date_to: q.setdefault("check_in_time", {})["$lte"] = date_to + "T23:59:59"
    if worker_id: q["worker_id"] = worker_id
    if site_id: q["site_id"] = site_id
    rows = []
    async for a in db.attendance_sessions.find(q, {"_id": 0}).sort("check_in_time", -1):
        ci = a.get("check_in_time", "")[:19].replace("T", " ")
        co = (a.get("check_out_time") or "")[:19].replace("T", " ") if a.get("check_out_time") else "-"
        rows.append([a.get("worker_name"), a.get("site_name"), ci, co,
                     f"{a.get('check_in_latitude', 0):.5f}, {a.get('check_in_longitude', 0):.5f}"])
    header = ["Worker", "Site", "Check-In", "Check-Out", "Check-In GPS"]
    return header, rows, None, "Attendance Report"


async def _rpt_site_installations(date_from, date_to, site_id):
    q = {"approval_status": "Approved"}
    if date_from: q.setdefault("work_date", {})["$gte"] = date_from
    if date_to: q.setdefault("work_date", {})["$lte"] = date_to
    if site_id: q["site_id"] = site_id
    by_site: dict = {}
    async for r in db.work_reports.find(q, {"_id": 0}):
        sid = r["site_id"]
        if sid not in by_site:
            by_site[sid] = {"site": r.get("site_name", "-"), "doors": 0, "amount": 0.0, "visits": 0}
        by_site[sid]["doors"] += r.get("door_count", 0)
        by_site[sid]["amount"] += float(r.get("total_amount", 0))
        by_site[sid]["visits"] += 1
    rows, td, ta = [], 0, 0.0
    for v in by_site.values():
        rows.append([v["site"], v["visits"], v["doors"], _rupees(v["amount"])])
        td += v["doors"]; ta += v["amount"]
    return ["Site", "Visits", "Doors", "Amount"], rows, ["TOTAL", "", td, _rupees(ta)], "Site-wise Installations Report"


async def _rpt_worker_performance(date_from, date_to, worker_id):
    q_reports = {}
    if date_from: q_reports.setdefault("work_date", {})["$gte"] = date_from
    if date_to: q_reports.setdefault("work_date", {})["$lte"] = date_to
    if worker_id: q_reports["worker_id"] = worker_id
    by_worker: dict = {}
    async for r in db.work_reports.find(q_reports, {"_id": 0}):
        wid = r["worker_id"]
        w = by_worker.setdefault(wid, {"name": r.get("worker_name", "-"), "doors": 0, "approved_amount": 0.0,
                                       "approved_reports": 0, "rejected_reports": 0, "correction_reports": 0,
                                       "days": set(), "sites": set()})
        w["days"].add(r.get("work_date"))
        w["sites"].add(r.get("site_id"))
        st = r.get("approval_status")
        if st == "Approved":
            w["approved_reports"] += 1
            w["doors"] += r.get("door_count", 0)
            w["approved_amount"] += float(r.get("total_amount", 0))
        elif st == "Rejected":
            w["rejected_reports"] += 1
        elif st == "Correction":
            w["correction_reports"] += 1

    rows = []
    for w in by_worker.values():
        days = len(w["days"])
        avg = round(w["doors"] / days, 2) if days else 0
        rows.append([w["name"], days, len(w["sites"]), w["doors"], avg,
                     _rupees(w["approved_amount"]), w["approved_reports"], w["rejected_reports"], w["correction_reports"]])
    header = ["Worker", "Days", "Sites", "Doors", "Avg/Day", "Approved Amount", "Approved", "Rejected", "Correction"]
    return header, rows, None, "Worker Performance Report"


async def _rpt_attendance_summary(date_from, date_to, worker_id, site_id):
    q = {}
    if date_from: q.setdefault("check_in_time", {})["$gte"] = date_from
    if date_to: q.setdefault("check_in_time", {})["$lte"] = date_to + "T23:59:59"
    if worker_id: q["worker_id"] = worker_id
    if site_id: q["site_id"] = site_id
    by_worker: dict = {}
    async for a in db.attendance_sessions.find(q, {"_id": 0}):
        wid = a["worker_id"]
        w = by_worker.setdefault(wid, {"name": a.get("worker_name", "-"), "sessions": 0, "sites": set(),
                                       "days": set(), "total_hours": 0.0})
        w["sessions"] += 1
        w["sites"].add(a.get("site_id"))
        w["days"].add((a.get("check_in_time") or "")[:10])
        if a.get("check_out_time"):
            try:
                ci = datetime.fromisoformat(a["check_in_time"].replace("Z", "+00:00"))
                co = datetime.fromisoformat(a["check_out_time"].replace("Z", "+00:00"))
                hrs = (co - ci).total_seconds() / 3600.0
                if hrs > 0: w["total_hours"] += hrs
            except Exception:
                pass
    rows = []
    for w in by_worker.values():
        rows.append([w["name"], len(w["days"]), w["sessions"], len(w["sites"]), f"{w['total_hours']:.1f}"])
    return ["Worker", "Days", "Sessions", "Sites", "Total Hours"], rows, None, "Attendance Summary Report"


REPORTS = {
    "worker_earnings": _rpt_worker_earnings,
    "payments": _rpt_payments,
    "attendance": _rpt_attendance,
    "site_installations": _rpt_site_installations,
    "worker_performance": _rpt_worker_performance,
    "attendance_summary": _rpt_attendance_summary,
}


async def _resolve_report(key: str, filters: dict):
    if key not in REPORTS:
        raise HTTPException(404, "Unknown report")
    fn = REPORTS[key]
    if key == "worker_earnings":
        return await fn(filters.get("date_from"), filters.get("date_to"), filters.get("worker_id"))
    if key == "payments":
        return await fn(filters.get("date_from"), filters.get("date_to"), filters.get("worker_id"))
    if key == "attendance":
        return await fn(filters.get("date_from"), filters.get("date_to"), filters.get("worker_id"), filters.get("site_id"))
    if key == "site_installations":
        return await fn(filters.get("date_from"), filters.get("date_to"), filters.get("site_id"))
    if key == "worker_performance":
        return await fn(filters.get("date_from"), filters.get("date_to"), filters.get("worker_id"))
    if key == "attendance_summary":
        return await fn(filters.get("date_from"), filters.get("date_to"), filters.get("worker_id"), filters.get("site_id"))
    raise HTTPException(400, "Bad report")


def _download_token(payload: dict) -> str:
    payload = {**payload, "jti": secrets.token_hex(8), "exp": now_utc() + timedelta(minutes=5)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)


async def _verify_download_token(token: str) -> dict:
    try:
        p = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
    except jwt.PyJWTError:
        raise HTTPException(401, "Invalid or expired token")
    jti = p.get("jti")
    if not jti:
        raise HTTPException(401, "Invalid token")
    # One-time use: these tokens travel in a URL query string and can end up in
    # access logs / browser history, so a replay within the 5-minute window is
    # rejected even though the signature and expiry are still valid.
    prior = await db.used_download_tokens.find_one_and_update(
        {"id": jti},
        {"$setOnInsert": {"id": jti, "expires_at": now_utc() + timedelta(minutes=6)}},
        upsert=True,
    )
    if prior is not None:
        raise HTTPException(401, "Download link already used")
    return p


async def _verify_admin_download_token(token: str) -> dict:
    """Like _verify_download_token, but additionally rejects any token not
    minted by an Admin. Invoice PDFs and report exports must never accept a
    worker-scoped salary-slip token — that token is only meant to be valid
    for that one worker's own slip, not for company-wide data."""
    payload = await _verify_download_token(token)
    if payload.get("role") != Role.admin.value:
        raise HTTPException(403, "Forbidden")
    return payload


@api.post("/reports/download-token")
async def create_download_token(u=Depends(require_role(Role.admin))):
    """Short-lived download token used by the mobile client to open export URLs."""
    return {"token": _download_token({"sub": u["id"], "role": u["role"]})}


@api.post("/worker/salary-slip-token")
async def worker_salary_slip_token(u=Depends(require_role(Role.worker))):
    """Narrowly-scoped download token — only valid for this worker's own salary slip,
    unlike the admin token above which can open any report/invoice/slip."""
    return {"token": _download_token({"sub": u["id"], "role": u["role"], "worker_id": u["worker_id"]})}


async def _company_settings():
    s = await db.settings.find_one({"id": "global"}, {"_id": 0})
    return s or {"company_name": "JONAH ENTERPRISES"}


def _build_xlsx(title: str, header, rows, footer, meta) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws["A1"] = meta["company_name"]
    ws["A1"].font = Font(size=16, bold=True, color="423226")
    ws["A2"] = title
    ws["A2"].font = Font(size=13, bold=True)
    ws["A3"] = f"Generated: {now_utc().strftime('%Y-%m-%d %H:%M UTC')}"
    if meta.get("filters"):
        ws["A4"] = "Filters: " + meta["filters"]
    start = 6
    for i, h in enumerate(header):
        cell = ws.cell(row=start, column=i + 1, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="423226", end_color="423226", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(rows):
        for ci, val in enumerate(row):
            ws.cell(row=start + 1 + ri, column=ci + 1, value=val)
    if footer:
        for ci, val in enumerate(footer):
            cell = ws.cell(row=start + 1 + len(rows), column=ci + 1, value=val)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="EADDBC", end_color="EADDBC", fill_type="solid")
    for col_i in range(1, len(header) + 1):
        ws.column_dimensions[chr(64 + col_i)].width = 22
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()


def _build_pdf(title: str, header, rows, footer, meta) -> bytes:
    from reportlab.lib import colors as rlc
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Heading1"], textColor=rlc.HexColor("#423226"))
    story = [
        Paragraph(meta["company_name"], title_style),
        Paragraph(title, styles["Heading2"]),
        Paragraph(f"Generated: {now_utc().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]),
    ]
    if meta.get("filters"):
        story.append(Paragraph(f"Filters: {meta['filters']}", styles["Normal"]))
    story.append(Spacer(1, 10))
    table_data = [header] + [list(r) for r in rows] + ([list(footer)] if footer else [])
    t = Table(table_data, repeatRows=1)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rlc.HexColor("#423226")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rlc.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rlc.HexColor("#FAF9F6"), rlc.white]),
        ("GRID", (0, 0), (-1, -1), 0.25, rlc.HexColor("#CFC9BC")),
    ])
    if footer:
        style.add("BACKGROUND", (0, -1), (-1, -1), rlc.HexColor("#EADDBC"))
        style.add("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")
    t.setStyle(style)
    story.append(t)
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


@api.get("/reports/export")
async def export_report(
    report: str = Query(...),
    fmt: str = Query("xlsx", pattern="^(xlsx|pdf)$"),
    token: Optional[str] = None,  # short-lived download token (used by browser link)
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    worker_id: Optional[str] = None,
    site_id: Optional[str] = None,
):
    # Downloads authenticate via a short-lived ?token= (created by /reports/download-token)
    # rather than the Authorization header — needed because mobile browsers/openURL
    # can't easily attach custom headers.
    if not token:
        raise HTTPException(401, "Missing download token")
    await _verify_admin_download_token(token)

    header, rows, footer, title = await _resolve_report(report, {
        "date_from": date_from, "date_to": date_to, "worker_id": worker_id, "site_id": site_id,
    })
    filter_bits = []
    if date_from: filter_bits.append(f"from {date_from}")
    if date_to: filter_bits.append(f"to {date_to}")
    meta = {"company_name": (await _company_settings()).get("company_name", "JONAH ENTERPRISES"),
            "filters": ", ".join(filter_bits) if filter_bits else None}

    if fmt == "xlsx":
        data = _build_xlsx(title, header, rows, footer, meta)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"
    else:
        data = _build_pdf(title, header, rows, footer, meta)
        media = "application/pdf"
        ext = "pdf"
    fname = f"{report}_{now_utc().strftime('%Y%m%d_%H%M')}.{ext}"
    return StreamingResponse(io.BytesIO(data), media_type=media, headers={
        "Content-Disposition": f'attachment; filename="{fname}"'
    })


# --- Startup / seed ------------------------------------------------------
# --- Time-based reminders --------------------------------------------------
REMINDER_INTERVAL_SECONDS = 1800  # 30 min — reminders are hours-scale, no need to poll tighter
OPEN_SESSION_REMINDER_HOURS = 12  # a normal shift is a few hours; 12+ almost certainly forgotten


async def remind_open_sessions():
    """A worker still checked in this long almost certainly forgot to check
    out (or the app died in the background) — nudge them and their admin
    once. reminder_sent guards against re-notifying every tick."""
    cutoff = (now_utc() - timedelta(hours=OPEN_SESSION_REMINDER_HOURS)).isoformat()
    cursor = db.attendance_sessions.find(
        {"check_out_time": None, "check_in_time": {"$lt": cutoff}, "reminder_sent": {"$ne": True}},
        {"_id": 0},
    )
    async for sess in cursor:
        worker_user = await db.users.find_one({"worker_id": sess["worker_id"]}, {"_id": 0, "id": 1})
        if worker_user:
            await add_notification(
                worker_user["id"], "Still checked in",
                f"You've been checked in at {sess['site_name']} since {formatted_ist(sess['check_in_time'])}. Don't forget to check out.",
            )
        async for admin in db.users.find({"role": Role.admin.value}, {"_id": 0, "id": 1}):
            await add_notification(
                admin["id"], "Worker still checked in",
                f"{sess['worker_name']} has been checked in at {sess['site_name']} since {formatted_ist(sess['check_in_time'])} and hasn't checked out.",
            )
        await db.attendance_sessions.update_one({"id": sess["id"]}, {"$set": {"reminder_sent": True}})


async def remind_stale_pending_approvals():
    """One reminder per calendar day (IST), not per tick — a digest, not spam."""
    today = datetime.now(IST).date().isoformat()
    settings = await db.settings.find_one({"id": "global"}, {"_id": 0, "last_pending_reminder_date": 1}) or {}
    if settings.get("last_pending_reminder_date") == today:
        return
    cutoff = (now_utc() - timedelta(hours=24)).isoformat()
    count = await db.work_reports.count_documents({"approval_status": "Pending", "created_at": {"$lt": cutoff}})
    if not count:
        return  # nothing stale yet — leave last_pending_reminder_date unset so a later tick today can still fire
    async for admin in db.users.find({"role": Role.admin.value}, {"_id": 0, "id": 1}):
        await add_notification(
            admin["id"], "Pending approvals",
            f"{count} work report(s) have been waiting over a day for approval.",
        )
    await db.settings.update_one({"id": "global"}, {"$set": {"last_pending_reminder_date": today}}, upsert=True)


def formatted_ist(iso_str: str) -> str:
    dt = datetime.fromisoformat(iso_str)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(IST).strftime("%d %b, %I:%M %p")


async def reminder_loop():
    while True:
        try:
            await remind_open_sessions()
            await remind_stale_pending_approvals()
        except Exception as e:
            logger.warning(f"reminder loop error (non-blocking): {e}")
        await asyncio.sleep(REMINDER_INTERVAL_SECONDS)


@app.on_event("startup")
async def startup():
    await db.users.create_index("username", unique=True)
    await db.workers.create_index("employee_id", unique=True)
    # TTL index so used one-time download tokens self-clean shortly after expiry
    await db.used_download_tokens.create_index("expires_at", expireAfterSeconds=0)
    # TTL index for login throttle records — storage hygiene only; the rate-limit
    # check itself uses an explicit time-window filter since Mongo's background
    # TTL sweep can lag up to ~60s behind the actual expiry.
    await db.login_attempts.create_index("created_at", expireAfterSeconds=LOGIN_WINDOW_SECONDS)
    await db.login_attempts.create_index("username")
    # Use partial index so only actual string sync_ids are enforced unique
    # (sparse indexes do NOT skip null values — they only skip missing fields).
    for coll_name in ("attendance_sessions", "work_reports", "payments"):
        coll = db[coll_name]
        # Drop stale sparse index if present, then recreate as partial.
        try:
            for idx_name, idx_info in (await coll.index_information()).items():
                keys = [k for k, _ in idx_info.get("key", [])]
                if keys == ["client_sync_id"]:
                    await coll.drop_index(idx_name)
        except Exception:
            pass
        await coll.create_index(
            "client_sync_id",
            unique=True,
            partialFilterExpression={"client_sync_id": {"$type": "string"}},
        )

    # Compound indexes on the fields actually filtered/aggregated on constantly —
    # invisible at today's data volume, a full collection scan on every dashboard
    # load and report once history grows into the thousands of documents.
    await db.work_reports.create_index([("worker_id", 1), ("work_date", -1)])
    await db.work_reports.create_index([("site_id", 1), ("work_date", -1)])
    await db.work_reports.create_index("approval_status")
    await db.attendance_sessions.create_index([("worker_id", 1), ("check_in_time", -1)])
    await db.attendance_sessions.create_index([("site_id", 1), ("check_in_time", -1)])
    await db.attendance_sessions.create_index("check_out_time")
    # Standalone index: the dashboard's "checked in today"/"sites today" queries
    # filter on check_in_time alone, which can't use either compound index above
    # as a prefix (both lead with worker_id/site_id first).
    await db.attendance_sessions.create_index("check_in_time")
    await db.payments.create_index([("worker_id", 1), ("payment_date", -1)])
    await db.notifications.create_index([("user_id", 1), ("created_at", -1)])
    await db.leaves.create_index([("worker_id", 1), ("status", 1)])
    await db.payment_requests.create_index([("worker_id", 1), ("status", 1)])
    await db.report_photos.create_index("report_id")
    await db.work_reports.create_index("excel_sync_status")
    await db.payments.create_index("excel_sync_status")

    # Backfill Excel Integration system fields onto records created before this
    # feature existed — never touches anything else on the document.
    await db.workers.update_many({"excel_match_status": {"$exists": False}}, {"$set": {
        "excel_party_id": None, "excel_match_status": ExcelMatchStatus.unresolved.value,
        "automation_enabled": False, "first_valid_entry_at": None, "automation_started_at": None,
    }})
    await db.sites.update_many({"excel_match_status": {"$exists": False}}, {"$set": {
        "excel_site_id": None, "excel_match_status": ExcelMatchStatus.unresolved.value,
    }})
    await db.door_types.update_many({"excel_match_status": {"$exists": False}}, {"$set": {
        "excel_item_id": None, "excel_match_status": ExcelMatchStatus.unresolved.value,
        "excel_work_category": None, "excel_specification": None,
    }})

    # settings
    if not await db.settings.find_one({"id": "global"}):
        await db.settings.insert_one({
            "id": "global",
            "company_name": "JONAH ENTERPRISES",
            "address": "Bengaluru, India",
            "mobile": "+91 98765 43210",
            "email": "contact@jonahenterprises.in",
            "currency": "₹",
            "default_rate": 250.0,
        })

    # seed a default door type — required config, not demo clutter: without at
    # least one, workers can't submit any report at all (min_length=1 on doors[]).
    if await db.door_types.count_documents({}) == 0:
        await db.door_types.insert_one({
            "id": str(uuid.uuid4()),
            "name": "Standard Door",
            "default_rate": 250.0,
            "status": "Active",
            "created_at": now_utc().isoformat(),
        })
        logger.info("Seeded default door type: Standard Door @ 250.0")

    # seed admin — bootstrap account needed to log in at all. Never a hardcoded
    # password: use ADMIN_INITIAL_PASSWORD if provided, otherwise generate a
    # random one-time password and log it so it's never a guessable default.
    if not await db.users.find_one({"username": "admin"}):
        admin_pwd = os.environ.get("ADMIN_INITIAL_PASSWORD")
        generated = admin_pwd is None
        if generated:
            admin_pwd = secrets.token_urlsafe(12)
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "username": "admin",
            "password_hash": pwd_ctx.hash(admin_pwd),
            "role": Role.admin.value,
            "name": "Admin",
            "disabled": False,
            "created_at": now_utc().isoformat(),
        })
        if generated:
            logger.warning(f"Seeded admin user. Initial password: {admin_pwd}  (change it immediately via /auth/change-password)")
        else:
            logger.info("Seeded admin user with password from ADMIN_INITIAL_PASSWORD")

    if not SEED_DEMO_DATA:
        return

    # seed workers
    if await db.workers.count_documents({}) == 0:
        demo = [
            {"employee_id": "EMP001", "name": "Ravi Kumar", "mobile": "+91 90000 00001", "username": "ravi", "password": "ravi123", "default_rate": 250.0},
            {"employee_id": "EMP002", "name": "Suresh Patel", "mobile": "+91 90000 00002", "username": "suresh", "password": "suresh123", "default_rate": 275.0},
            {"employee_id": "EMP003", "name": "Vinod Sharma", "mobile": "+91 90000 00003", "username": "vinod", "password": "vinod123", "default_rate": 250.0},
        ]
        for w in demo:
            wid = str(uuid.uuid4())
            uid = str(uuid.uuid4())
            await db.users.insert_one({
                "id": uid,
                "username": w["username"],
                "password_hash": pwd_ctx.hash(w["password"]),
                "role": Role.worker.value,
                "name": w["name"],
                "worker_id": wid,
                "disabled": False,
                "created_at": now_utc().isoformat(),
            })
            await db.workers.insert_one({
                "id": wid,
                "user_id": uid,
                "employee_id": w["employee_id"],
                "name": w["name"],
                "mobile": w["mobile"],
                "username": w["username"],
                "joining_date": now_utc().date().isoformat(),
                "default_rate": w["default_rate"],
                "status": "Active",
                "notes": "",
                "created_at": now_utc().isoformat(),
            })
        logger.info("Seeded 3 workers")

    # seed sites
    if await db.sites.count_documents({}) == 0:
        for s in [
            {"site_name": "Prestige Lakeside", "client_name": "Prestige Group", "address": "Whitefield, Bengaluru", "latitude": 12.9698, "longitude": 77.7500, "contact_person": "Rahul", "contact_number": "+91 98111 11111"},
            {"site_name": "Brigade Meadows", "client_name": "Brigade Group", "address": "Kanakapura Rd, Bengaluru", "latitude": 12.8500, "longitude": 77.5500, "contact_person": "Anil", "contact_number": "+91 98111 22222"},
            {"site_name": "Sobha Dream Acres", "client_name": "Sobha Ltd", "address": "Panathur, Bengaluru", "latitude": 12.9400, "longitude": 77.7100, "contact_person": "Deepak", "contact_number": "+91 98111 33333"},
            {"site_name": "Godrej Woodsville", "client_name": "Godrej Properties", "address": "Hosur Rd, Bengaluru", "latitude": 12.8800, "longitude": 77.6600, "contact_person": "Meera", "contact_number": "+91 98111 44444"},
        ]:
            await db.sites.insert_one({**s, "id": str(uuid.uuid4()), "status": "Active", "notes": "", "created_at": now_utc().isoformat()})
        logger.info("Seeded 4 sites")

    # seed sample work reports + payments (minimal for demo)
    if await db.work_reports.count_documents({}) == 0:
        admin_user = await db.users.find_one({"role": Role.admin.value}, {"_id": 0})
        admin_id = admin_user["id"] if admin_user else "admin"
        workers = await db.workers.find({}, {"_id": 0}).to_list(10)
        sites = await db.sites.find({}, {"_id": 0}).to_list(10)
        if workers and sites:
            w = workers[0]
            s = sites[0]
            # closed session
            sess_id = str(uuid.uuid4())
            await db.attendance_sessions.insert_one({
                "id": sess_id, "worker_id": w["id"], "worker_name": w["name"],
                "site_id": s["id"], "site_name": s["site_name"],
                "check_in_time": (now_utc() - timedelta(days=2, hours=8)).isoformat(),
                "check_in_latitude": s["latitude"], "check_in_longitude": s["longitude"],
                "check_in_accuracy": 10.0,
                "check_out_time": (now_utc() - timedelta(days=2, hours=1)).isoformat(),
                "check_out_latitude": s["latitude"], "check_out_longitude": s["longitude"],
                "check_out_accuracy": 12.0, "work_completed": True, "remarks": "Completed 8 doors",
                "created_at": now_utc().isoformat(),
            })
            await db.work_reports.insert_one({
                "id": str(uuid.uuid4()), "session_id": sess_id,
                "worker_id": w["id"], "worker_name": w["name"],
                "site_id": s["id"], "site_name": s["site_name"],
                "work_date": (now_utc() - timedelta(days=2)).date().isoformat(),
                "door_count": 8, "rate_per_door": w["default_rate"],
                "total_amount": 8 * w["default_rate"],
                "notes": "All 8 main entrance doors installed",
                "work_completed": True,
                "before_photo_ids": [], "after_photo_ids": [],
                "approval_status": "Approved",
                "approval_remarks": "Verified, good work",
                "approved_by": admin_id, "approved_by_name": "Admin",
                "approved_at": (now_utc() - timedelta(days=1)).isoformat(),
                "created_at": (now_utc() - timedelta(days=2)).isoformat(),
            })
            await db.payments.insert_one({
                "id": str(uuid.uuid4()), "worker_id": w["id"], "worker_name": w["name"],
                "payment_date": (now_utc() - timedelta(days=1)).date().isoformat(),
                "amount": 1500.0, "payment_method": "UPI",
                "transaction_reference": "UPI/DEMO/001",
                "notes": "Partial advance",
                "created_by": admin_id,
                "created_at": now_utc().isoformat(),
            })
            logger.info("Seeded sample report & payment")

    app.state.reminder_task = asyncio.create_task(reminder_loop())
    app.state.excel_sync_task = asyncio.create_task(excel_sync_loop())


@app.on_event("shutdown")
async def shutdown():
    task = getattr(app.state, "reminder_task", None)
    if task:
        task.cancel()
    excel_task = getattr(app.state, "excel_sync_task", None)
    if excel_task:
        excel_task.cancel()
    client.close()


app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=False,
    allow_origins=CORS_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
)

FRONTEND_DIR = ROOT_DIR / "frontend"
if FRONTEND_DIR.is_dir():
    @app.middleware("http")
    async def no_cache_frontend(request, call_next):
        # Mobile Safari in particular will keep serving a stale cached JS bundle
        # long after the file changes on disk unless the server explicitly forbids it.
        response = await call_next(request)
        if not request.url.path.startswith("/api"):
            response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        return response

    app.mount("/", StaticFiles(directory=str(FRONTEND_DIR), html=True), name="frontend")
